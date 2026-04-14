import sqlite3
import datetime
import time
import threading
import firebase_admin
import plotly
import plotly.graph_objs as go
import pandas as pd
import json
import io
import os
from firebase_admin import credentials, db
from flask import Flask, request, render_template, redirect, url_for, jsonify, session, Response, send_file
from dateutil.relativedelta import relativedelta
from functools import wraps

# ================= KONFIGURASI =================
DB_NAME = "database_lab.sqlite"
# Ganti dengan URL Database Anda (Akhiri dengan tanda slash /)
FIREBASE_DB_URL = "https://absensi-lab-ap-default-rtdb.asia-southeast1.firebasedatabase.app/" 

# Password Global (Harus sama dengan di HTML)
ONLINE_TOKEN = "LAB_JAYA"
ADMIN_PASSWORD = "ServerAbsensiLABSESL"

app = Flask(__name__)
app.secret_key = "kunci_rahasia_sesl_lab_sangat_rahasia"

# --- GLOBAL TEMPLATE CONTEXT ---
@app.context_processor
def inject_globals():
    global CACHED_JADWAL
    raw = [j.get('lokasi_lab') for j in CACHED_JADWAL if j.get('lokasi_lab')]
    labs_nav = sorted(list(set([str(l).strip() for l in raw])))
    path = request.path
    ap = 'home'
    al = ''
    if path.startswith('/active'): ap = 'active'
    elif path.startswith('/log'): ap = 'log'
    elif path.startswith('/rekap'):
        ap = 'rekap'
        parts = path.split('/')
        if len(parts) >= 3: al = parts[2]
    elif path.startswith('/admin'): ap = 'admin'
    elif path.startswith('/profil'): ap = 'rekap'
    return dict(all_labs_for_nav=labs_nav, active_page=ap, active_lab=al)

# Cache Data
CACHED_JADWAL = [] 
CACHED_DEVICE_CONTROL = {}
IS_MASTER_DATA_READY = False

# ================= SETUP FIREBASE =================
cred = credentials.Certificate("firebase_credentials.json")
firebase_admin.initialize_app(cred, {'databaseURL': FIREBASE_DB_URL})


# --- FORMAT WAKTU STANDAR UNTUK DATABASE ---
# Kita gunakan format ini di semua tempat agar konsisten
DB_TIME_FORMAT = "%Y-%m-%d %H:%M:%S"

# ================= FUNGSI DATABASE =================
def init_db():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    
    # Tabel Users
    cursor.execute('''CREATE TABLE IF NOT EXISTS users (
        fingerprint_id INTEGER PRIMARY KEY, 
        nama TEXT, 
        id_asisten_kampus TEXT,
        hak_akses TEXT,
        is_synced INTEGER DEFAULT 1,
        sync_action TEXT DEFAULT NULL
    )''')
    
    # Tabel Logs
    cursor.execute('''CREATE TABLE IF NOT EXISTS logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        firebase_key TEXT,
        fingerprint_id INTEGER, nama TEXT, id_asisten_kampus TEXT,
        waktu_masuk DATETIME, waktu_keluar DATETIME,
        status TEXT, lokasi_lab TEXT, kelas TEXT,
        is_synced INTEGER DEFAULT 0 
    )''')

    # Tabel Active Sessions
    cursor.execute('''CREATE TABLE IF NOT EXISTS active_sessions (
        fingerprint_id INTEGER PRIMARY KEY,
        nama TEXT, id_asisten_kampus TEXT,
        waktu_masuk DATETIME,
        jam_selesai_kelas DATETIME,
        lokasi_lab TEXT, kelas TEXT,
        log_db_id INTEGER
    )''')

    #Tabel Jadwal Kelas
    cursor.execute('''CREATE TABLE IF NOT EXISTS schedules (
        id_jadwal TEXT PRIMARY KEY,
        nama_kelas TEXT,
        hari TEXT,
        jam_mulai TEXT,
        jam_selesai TEXT,
        lokasi_lab TEXT,
        is_online INTEGER,
        is_synced INTEGER DEFAULT 1,
        sync_action TEXT DEFAULT NULL
    )''')
    
    conn.commit(); conn.close()

def hitung_durasi_menit(waktu_masuk_str, waktu_keluar_str):
    if not waktu_masuk_str or not waktu_keluar_str:
        return 0
    try:
        masuk = datetime.datetime.strptime(waktu_masuk_str, DB_TIME_FORMAT)
        keluar = datetime.datetime.strptime(waktu_keluar_str, DB_TIME_FORMAT)
        if keluar < masuk:
            return 0
        durasi_total = keluar - masuk
        
        total_menit_genap = durasi_total.total_seconds() // 60
        return total_menit_genap
        
    except:
        return 0

def format_rupiah(angka):
    return f"Rp{angka:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

# ================= KEAMANAN & LOGIN =================
def login_required(f):
    """Decorator untuk mengunci halaman agar hanya bisa diakses admin"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('is_admin'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        if request.form['password'] == ADMIN_PASSWORD:
            session['is_admin'] = True
            return redirect(url_for('dashboard_home'))
        else:
            error = 'Password salah!'
            
    # [HTML REMOVED - Moved to templates/]
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.pop('is_admin', None)
    return redirect(url_for('dashboard_home'))

# ==============================================================================
# API ENDPOINT KHUSUS ADMIN (READ & WRITE)
# ==============================================================================

# --- 1. API READ (MENGAMBIL DATA UNTUK DITAMPILKAN) ---
@app.route('/api/admin/asisten')
@login_required
def api_admin_asisten():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM users WHERE sync_action != 'DELETE' OR sync_action IS NULL ORDER BY fingerprint_id ASC")
    asisten =[dict(row) for row in cur.fetchall()]
    conn.close()
    return jsonify({'asisten': asisten})

@app.route('/api/admin/jadwal')
@login_required
def api_admin_jadwal():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM schedules WHERE sync_action != 'DELETE' OR sync_action IS NULL ORDER BY lokasi_lab ASC, hari ASC")
    jadwal =[dict(row) for row in cur.fetchall()]
    conn.close()
    return jsonify({'jadwal': jadwal})

# ---> INI RUTE YANG SEBELUMNYA HILANG/TERTIMPA (Penyebab Error 415) <---
@app.route('/api/admin/device')
@login_required
def api_admin_device():
    global CACHED_DEVICE_CONTROL, CACHED_JADWAL
    raw_labs =[jadwal.get('lokasi_lab') for jadwal in CACHED_JADWAL if jadwal.get('lokasi_lab')]
    all_labs = sorted(list(set([str(lab).strip() for lab in raw_labs])))
    
    return jsonify({
        'device_control': CACHED_DEVICE_CONTROL,
        'labs': all_labs
    })

# --- 2. API WRITE (MENGUBAH DATA / ACTION) ---
@app.route('/api/admin/device/update', methods=['POST'])
@login_required
def api_admin_device_update():
    data = request.json
    lab_name = data.get('lab_name')
    mode = data.get('mode')
    target_id = int(data.get('target_id', 0))

    global CACHED_DEVICE_CONTROL
    if lab_name not in CACHED_DEVICE_CONTROL:
        CACHED_DEVICE_CONTROL[lab_name] = {}
        
    # UPDATE RAM LOKAL
    CACHED_DEVICE_CONTROL[lab_name]['mode'] = mode
    CACHED_DEVICE_CONTROL[lab_name]['target_id'] = target_id

    # UPDATE FIREBASE DI BACKGROUND
    def sync_device_to_firebase():
        try:
            db.reference(f'device_control/{lab_name}').update({
                'mode': mode, 'target_id': target_id
            })
        except Exception as e:
            print(f"[DEVICE-SYNC ERROR] {e}")
            
    threading.Thread(target=sync_device_to_firebase).start()
    return jsonify({'status': 'success'})

@app.route('/api/admin/asisten/action', methods=['POST', 'DELETE'])
@login_required
def api_admin_asisten_action():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    if request.method == 'POST':
        data = request.json
        action = data.get('action')
        fid = int(data.get('fingerprint_id'))
        nama = data.get('nama')
        id_kampus = data.get('id_asisten_kampus')
        hak_akses = data.get('hak_akses')
        try:
            if action == 'INSERT':
                cur.execute('''INSERT INTO users (fingerprint_id, nama, id_asisten_kampus, hak_akses, is_synced, sync_action)
                               VALUES (?, ?, ?, ?, 0, 'INSERT')''', (fid, nama, id_kampus, hak_akses))
            elif action == 'UPDATE':
                old_fid = int(data.get('old_fingerprint_id'))
                
                # --- PERBAIKAN: CEK PERUBAHAN ID KAMPUS (FIREBASE KEY) ---
                cur.execute("SELECT id_asisten_kampus FROM users WHERE fingerprint_id=?", (old_fid,))
                old_row = cur.fetchone()
                if old_row and old_row[0] != id_kampus:
                    # Jika ID berubah, perintahkan fungsi background untuk menghapus node lama di Firebase
                    threading.Thread(target=safe_delete_firebase_node, args=(f'asisten_master/{old_row[0]}',), daemon=True).start()
                # ---------------------------------------------------------

                cur.execute('''UPDATE users SET fingerprint_id=?, nama=?, id_asisten_kampus=?, hak_akses=?, is_synced=0, sync_action='UPDATE'
                               WHERE fingerprint_id=?''', (fid, nama, id_kampus, hak_akses, old_fid))
            conn.commit()
        except sqlite3.IntegrityError:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Fingerprint ID atau ID Kampus sudah dipakai.'}), 400
    elif request.method == 'DELETE':
        fid = int(request.json.get('fingerprint_id'))
        cur.execute("UPDATE users SET is_synced=0, sync_action='DELETE' WHERE fingerprint_id=?", (fid,))
        conn.commit()
    conn.close()
    return jsonify({'status': 'success'})

@app.route('/api/admin/jadwal/action', methods=['POST', 'DELETE'])
@login_required
def api_admin_jadwal_action():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    if request.method == 'POST':
        data = request.json
        action = data.get('action')
        
        # --- PERBAIKAN: GENERATE ID JADWAL SECARA DINAMIS ---
        nama_k = data.get('nama_kelas')
        lokasi = data.get('lokasi_lab')
        
        # Buat ID berdasarkan Nama Lab dan Kelas yang diinput (misal: LAB_AP & TTLA -> AP_TTLA)
        kode_lab = lokasi.replace("LAB_", "")
        kode_kelas = nama_k.replace(" ", "")
        generated_id = f"{kode_lab}_{kode_kelas}".upper()
        # ----------------------------------------------------

        hari = data.get('hari')
        jm = data.get('jam_mulai')
        js = data.get('jam_selesai')
        is_on = int(data.get('is_online'))
        
        if action == 'INSERT':
            try:
                cur.execute('''INSERT INTO schedules (id_jadwal, nama_kelas, hari, jam_mulai, jam_selesai, lokasi_lab, is_online, is_synced, sync_action)
                               VALUES (?, ?, ?, ?, ?, ?, ?, 0, 'INSERT')''', (generated_id, nama_k, hari, jm, js, lokasi, is_on))
            except sqlite3.IntegrityError:
                conn.close()
                return jsonify({'status': 'error', 'message': 'Jadwal Kelas ini di Lab tersebut sudah ada!'}), 400
                
        elif action == 'UPDATE':
            old_id_j = data.get('id_jadwal')
            
            # --- PERBAIKAN: CEK PERUBAHAN NAMA LAB / KELAS (FIREBASE KEY) ---
            if old_id_j != generated_id:
                # Jika Key berubah, hapus node lama di Firebase secara background
                threading.Thread(target=safe_delete_firebase_node, args=(f'jadwal_kelas/{old_id_j}',), daemon=True).start()
            # ----------------------------------------------------------------

            # Lakukan update, pastikan id_jadwal (Primary Key) di SQLite juga ikut terupdate
            cur.execute('''UPDATE schedules SET id_jadwal=?, nama_kelas=?, hari=?, jam_mulai=?, jam_selesai=?, lokasi_lab=?, is_online=?, is_synced=0, sync_action='UPDATE'
                           WHERE id_jadwal=?''', (generated_id, nama_k, hari, jm, js, lokasi, is_on, old_id_j))
        conn.commit()
    elif request.method == 'DELETE':
        id_j = request.json.get('id_jadwal')
        cur.execute("UPDATE schedules SET is_synced=0, sync_action='DELETE' WHERE id_jadwal=?", (id_j,))
        conn.commit()
    conn.close()
    
    # Refresh cache RAM seketika
    load_jadwal_lokal() 
    return jsonify({'status': 'success'})

# ==============================================================================
# API ENDPOINT KHUSUS ADMIN (LOG ACTION)
# ==============================================================================
@app.route('/api/admin/log/action', methods=['POST'])
@login_required
def api_admin_log_action():
    conn = sqlite3.connect(DB_NAME)
    try:
        cur = conn.cursor()
        data = request.json
        action = data.get('action')
        
        log_id = data.get('id')
        nama = data.get('nama')
        id_kampus = data.get('id_asisten_kampus')
        lokasi = data.get('lokasi_lab')
        kelas = data.get('kelas')
        status = data.get('status')
        
        # Gabungkan Tanggal dan Jam
        waktu_masuk = f"{data.get('in_date')} {data.get('in_time')}" if data.get('in_date') and data.get('in_time') else None
        waktu_keluar = f"{data.get('out_date')} {data.get('out_time')}" if data.get('out_date') and data.get('out_time') else None

        # Cari Fingerprint ID dari master users
        cur.execute("SELECT fingerprint_id FROM users WHERE id_asisten_kampus=?", (id_kampus,))
        user_row = cur.fetchone()
        fp_id = user_row[0] if user_row else 0

        # Helper: Cari estimasi jam selesai kelas dari Master Jadwal
        def get_jam_selesai(lab, kls, wt_masuk):
            if not wt_masuk: return None
            global CACHED_JADWAL
            for jdwl in CACHED_JADWAL:
                if jdwl.get('lokasi_lab') == lab and jdwl.get('nama_kelas') == kls:
                    try:
                        mdt = datetime.datetime.strptime(wt_masuk, DB_TIME_FORMAT)
                        j_h, j_m = map(int, jdwl['jam_selesai'].split(':'))
                        return mdt.replace(hour=j_h, minute=j_m, second=0).strftime(DB_TIME_FORMAT)
                    except: pass
            # Default jika tidak ada jadwal cocok: 3 Jam dari waktu masuk
            mdt = datetime.datetime.strptime(wt_masuk, DB_TIME_FORMAT)
            return (mdt + datetime.timedelta(hours=3)).strftime(DB_TIME_FORMAT)

        if action == 'INSERT':
            cur.execute('''INSERT INTO logs (fingerprint_id, nama, id_asisten_kampus, waktu_masuk, waktu_keluar, status, lokasi_lab, kelas, is_synced)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0)''', 
                           (fp_id, nama, id_kampus, waktu_masuk, waktu_keluar, status, lokasi, kelas))
            last_id = cur.lastrowid
            
            # Jika status MASUK dan terjadi hari ini, masukkan juga ke Active Sessions
            if status.upper() == 'MASUK' and waktu_masuk:
                if waktu_masuk[:10] == datetime.date.today().strftime('%Y-%m-%d'):
                    jam_selesai = get_jam_selesai(lokasi, kelas, waktu_masuk)
                    cur.execute('''INSERT INTO active_sessions (fingerprint_id, nama, id_asisten_kampus, waktu_masuk, jam_selesai_kelas, lokasi_lab, kelas, log_db_id)
                                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                                   (fp_id, nama, id_kampus, waktu_masuk, jam_selesai, lokasi, kelas, last_id))

        elif action == 'UPDATE':
            cur.execute('''UPDATE logs SET nama=?, id_asisten_kampus=?, waktu_masuk=?, waktu_keluar=?, status=?, lokasi_lab=?, kelas=?, is_synced=0
                           WHERE id=?''', (nama, id_kampus, waktu_masuk, waktu_keluar, status, lokasi, kelas, log_id))
            
            # Sinkronisasi ke Active Sessions jika log diubah
            if status.upper() == 'MASUK' and waktu_masuk:
                jam_selesai = get_jam_selesai(lokasi, kelas, waktu_masuk)
                cur.execute("SELECT * FROM active_sessions WHERE log_db_id=?", (log_id,))
                if cur.fetchone():
                    cur.execute("UPDATE active_sessions SET waktu_masuk=?, lokasi_lab=?, kelas=?, jam_selesai_kelas=? WHERE log_db_id=?",
                                (waktu_masuk, lokasi, kelas, jam_selesai, log_id))
                elif waktu_masuk[:10] == datetime.date.today().strftime('%Y-%m-%d'):
                    cur.execute('''INSERT INTO active_sessions (fingerprint_id, nama, id_asisten_kampus, waktu_masuk, jam_selesai_kelas, lokasi_lab, kelas, log_db_id)
                                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                                   (fp_id, nama, id_kampus, waktu_masuk, jam_selesai, lokasi, kelas, log_id))
            else:
                # Jika diubah jadi KELUAR, pastikan terhapus dari active_session
                cur.execute("DELETE FROM active_sessions WHERE log_db_id=?", (log_id,))

        conn.commit()
        return jsonify({'status': 'success'})
    except Exception as e:
        print(f"[ADMIN LOG ERROR] {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

# ==============================================================================
# ROUTE DASHBOARD ADMIN (UI & JAVASCRIPT)
# ==============================================================================
@app.route('/admin')
@login_required
def dashboard_admin():
    global CACHED_JADWAL
    raw_labs =[jadwal.get('lokasi_lab') for jadwal in CACHED_JADWAL if jadwal.get('lokasi_lab')]
    all_labs_for_nav = sorted(list(set([str(lab).strip() for lab in raw_labs])))

    # [HTML REMOVED - Moved to templates/]
    return render_template('admin_panel.html')

# ================= LOGIKA JADWAL =================
def get_current_schedule(lab_name, check_time=None):
    """
    Mencari jadwal aktif.
    check_time: Objek datetime. Jika None, gunakan waktu sekarang.
    """
    if check_time is None:
        target_waktu = datetime.datetime.now()
    else:
        target_waktu = check_time
        
    hari_map = {
        "Monday": "Senin", "Tuesday": "Selasa", "Wednesday": "Rabu", 
        "Thursday": "Kamis", "Friday": "Jumat", "Saturday": "Sabtu", "Sunday": "Minggu"
    }
    hari_ini = hari_map.get(target_waktu.strftime("%A"), "Unknown")
    
    jam_target_menit = target_waktu.hour * 60 + target_waktu.minute

    # print(f"[DEBUG JADWAL] Cek {lab_name}, Hari {hari_ini}, Jam {jam_target_menit}")

    for jadwal in CACHED_JADWAL:
        # Filter Lokasi & Hari
        if jadwal.get('lokasi_lab') == lab_name and jadwal.get('hari', '').lower() == hari_ini.lower():
            
            # Khusus Request Online: Pastikan jadwal ini bertanda is_online=True
            # (Validasi ini bisa diperketat di caller, tapi di sini kita filter umum dulu)
            
            try:
                jm_h, jm_m = map(int, jadwal.get('jam_mulai', '00:00').split(':'))
                js_h, js_m = map(int, jadwal.get('jam_selesai', '00:00').split(':'))
                
                start_menit = jm_h * 60 + jm_m
                end_menit = js_h * 60 + js_m
                
                # Toleransi 15 menit sebelum
                if (jam_target_menit >= (start_menit - 15)) and (jam_target_menit < end_menit):
                    waktu_mulai_obj = target_waktu.replace(hour=jm_h, minute=jm_m, second=0, microsecond=0)
                    waktu_selesai_obj = target_waktu.replace(hour=js_h, minute=js_m, second=0, microsecond=0)
                    return jadwal, waktu_mulai_obj, waktu_selesai_obj
            except Exception as e:
                continue
                
    return None, None, None

# ================= BACKGROUND THREADS =================
def sync_asisten_from_firebase():
    """Menarik data asisten dari Firebase, AMAN dari menimpa data antrean lokal"""
    try:
        data_asisten = db.reference('asisten_master').get()
        conn = sqlite3.connect(DB_NAME, timeout=10) 
        cur = conn.cursor()
        
        # 1. Ambil ID asisten yang sedang antre upload/hapus di lokal
        cur.execute("SELECT id_asisten_kampus FROM users WHERE is_synced=0")
        unsynced_ids = [row[0] for row in cur.fetchall()]
        
        if data_asisten is not None:
            # 2. Hapus HANYA asisten yang sudah sinkron (aman ditimpa)
            cur.execute("DELETE FROM users WHERE is_synced=1")
            
            for key, val in data_asisten.items():
                # 3. JANGAN TIMPA JIKA ID INI ADA DI ANTREAN LOKAL
                if key not in unsynced_ids:
                    fid = val.get('fingerprint_id', 0)
                    if fid > 0:
                        hak_akses = val.get('hak_akses', '')
                        cur.execute('''INSERT INTO users 
                                    (fingerprint_id, nama, id_asisten_kampus, hak_akses, is_synced, sync_action) 
                                    VALUES (?, ?, ?, ?, 1, NULL)''', 
                                   (fid, val.get('nama', 'Unknown'), val.get('id_asisten_kampus', key), hak_akses)) 
        else:
            cur.execute("DELETE FROM users WHERE is_synced=1")
            
        conn.commit(); conn.close()
        print("[MASTER-SYNC] Data Asisten diperbarui (Aman dari timpaan).")
    except Exception as e:
        print(f"[MASTER-SYNC ERROR] Gagal sync asisten: {e}")
        try: conn.close() 
        except: pass

def load_jadwal_lokal():
    """Memuat jadwal dari SQLite ke RAM (CACHED_JADWAL) saat Server Startup"""
    global CACHED_JADWAL, IS_MASTER_DATA_READY
    try:
        conn = sqlite3.connect(DB_NAME)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT * FROM schedules WHERE sync_action != 'DELETE' OR sync_action IS NULL")
        rows = cur.fetchall()
        
        temp_jadwal =[]
        for row in rows:
            jadwal_dict = dict(row)
            # Konversi integer 1/0 kembali menjadi boolean True/False untuk logika Python
            jadwal_dict['is_online'] = bool(jadwal_dict['is_online'])
            temp_jadwal.append(jadwal_dict)
            
        CACHED_JADWAL = temp_jadwal
        conn.close()
        
        # KARENA DATA SUDAH ADA DI LOKAL, KITA BISA LANGSUNG ANGKAT BENDERA!
        IS_MASTER_DATA_READY = True
        print(f"[STARTUP-JADWAL] Jadwal dimuat dari DB Lokal: {len(CACHED_JADWAL)} kelas. Master Data READY.")
    except Exception as e:
        print(f"[STARTUP-JADWAL ERROR] Gagal muat jadwal lokal: {e}")

def sync_jadwal_from_firebase():
    """Menarik data jadwal dari Firebase, AMAN dari menimpa data antrean lokal"""
    global CACHED_JADWAL
    try:
        data_jadwal = db.reference('jadwal_kelas').get()
        conn = sqlite3.connect(DB_NAME, timeout=10)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        
        cur.execute("SELECT id_jadwal FROM schedules WHERE is_synced=0")
        unsynced_ids = [row[0] for row in cur.fetchall()]
        
        if data_jadwal is not None:
            cur.execute("DELETE FROM schedules WHERE is_synced=1")
            temp_jadwal =[]
            
            # Masukkan data antrean lokal ke CACHE RAM agar tidak hilang sesaat
            cur.execute("SELECT * FROM schedules WHERE is_synced=0 AND sync_action != 'DELETE'")
            for local_row in cur.fetchall():
                jd_dict = dict(local_row)
                jd_dict['is_online'] = bool(jd_dict['is_online'])
                temp_jadwal.append(jd_dict)
            
            for key, val in data_jadwal.items(): 
                if key not in unsynced_ids:
                    is_online_int = 1 if val.get('is_online', False) else 0
                    cur.execute('''INSERT INTO schedules 
                        (id_jadwal, nama_kelas, hari, jam_mulai, jam_selesai, lokasi_lab, is_online, is_synced, sync_action)
                        VALUES (?, ?, ?, ?, ?, ?, ?, 1, NULL)''',
                        (key, val.get('nama_kelas'), val.get('hari'), val.get('jam_mulai'), 
                         val.get('jam_selesai'), val.get('lokasi_lab'), is_online_int))
                    temp_jadwal.append(val)
                
            conn.commit()
            CACHED_JADWAL = temp_jadwal
            print(f"[MASTER-SYNC] Jadwal diperbarui. Total: {len(CACHED_JADWAL)} kelas aktif.")
        else:
            cur.execute("DELETE FROM schedules WHERE is_synced=1")
            conn.commit()
            
        conn.close()
    except Exception as e:
        print(f"[MASTER-SYNC ERROR] Gagal sync jadwal: {e}")
        try: conn.close() 
        except: pass

def stream_master_data_listener():
    """
    Mendengarkan perubahan Real-Time pada data asisten dan jadwal.
    Jika ada perubahan (tambah/edit/hapus), langsung tarik data terbaru.
    """
    # Event Handler
    def listener_asisten(event):
        # Abaikan event awal karena sudah di-handle oleh initial load
        if event.path == "/": return
        sync_asisten_from_firebase()

    def listener_jadwal(event):
        if event.path == "/": return
        sync_jadwal_from_firebase()

    while True:
        print("[MASTER-STREAM] Memulai koneksi Listener Master Data...")
        try:
            # Jalankan initial load sekali saat koneksi pertama kali terbuka
            sync_asisten_from_firebase()
            sync_jadwal_from_firebase()
            # Buka telinga (Listener)
            stream_a = db.reference('asisten_master').listen(listener_asisten)
            stream_j = db.reference('jadwal_kelas').listen(listener_jadwal)
            
            while True:
                time.sleep(60) # Tahan thread agar tidak mati
                
        except Exception as e:
            print(f"[MASTER-STREAM KONEKSI PUTUS/ERROR] {e}. Me-restart dalam 10 detik...")
            try: stream_a.close()
            except: pass
            try: stream_j.close()
            except: pass
            time.sleep(10)

# ================= BACKGROUND THREADS (MASTER DATA WATCHDOG) =================
def task_master_data_watchdog():
    """
    Patroli rutin setiap 10 menit (600 detik).
    Mengatasi kasus langka di mana koneksi internet putus diam-diam 
    sehingga listener tidak menyadari adanya perubahan data di Firebase.
    """
    time.sleep(60)
    print("[MASTER-WATCHDOG] Patroli Data Master Berjalan...")
    while True:
        # Kita panggil saja fungsi helper-nya
        sync_asisten_from_firebase()
        sync_jadwal_from_firebase()
        
        # Tidur 10 Menit
        time.sleep(600)

# ================= BACKGROUND THREADS (FINAL AUTO TAP-OUT & CLEANUP MERGE - FIX) =================
def task_smart_auto_tap_out():
    """
    Tugas cerdas yang berjalan terus menerus untuk:
    1. Menutup sesi yang jadwalnya sudah habis (normal).
    2. Membersihkan sesi basi dari hari sebelumnya (fungsi cleanup).
    """
    print("[SMART-AUTO] Thread Auto Tap-Out & Cleanup Berjalan...")
    while True:
        try:
            conn = sqlite3.connect(DB_NAME)
            cur = conn.cursor()
            now = datetime.datetime.now()
            today_str = datetime.date.today().strftime('%Y-%m-%d')
            
            cur.execute("""
                SELECT fingerprint_id, nama, log_db_id, jam_selesai_kelas, waktu_masuk 
                FROM active_sessions 
                WHERE ? >= jam_selesai_kelas OR date(waktu_masuk) != ?
            """, (now, today_str))
            expired_sessions = cur.fetchall()
            
            for sess in expired_sessions:
                fp_id, nama_user, log_db_id, jam_selesai_str, waktu_masuk_str = sess
                
                status_final = 'Keluar (Auto)'
                
                waktu_masuk_tanggal_obj = datetime.datetime.strptime(waktu_masuk_str, DB_TIME_FORMAT).date()
                
                if waktu_masuk_tanggal_obj != datetime.date.today():
                    status_final = 'Keluar (Auto-Cleanup)'

                print(f"[SMART-AUTO] {nama_user} Auto Tap-Out ({status_final}).")

                waktu_selesai_kelas = datetime.datetime.strptime(jam_selesai_str, DB_TIME_FORMAT)
                
                cur.execute("UPDATE logs SET status=?, waktu_keluar=?, is_synced=0 WHERE id=?", 
                            (status_final, waktu_selesai_kelas.strftime(DB_TIME_FORMAT), log_db_id))
                
                cur.execute("DELETE FROM active_sessions WHERE fingerprint_id=?", (fp_id,))
                
            conn.commit(); conn.close()
        except Exception as e:
            print(f"[SMART-AUTO ERROR] {e}")
        time.sleep(30)

# ================= BACKGROUND THREADS (FINAL TIMESTAMP FIX) =================
def task_upload_logs():
    print("[UPLOAD] Thread Upload Logs Berjalan...")
    last_used_timestamp = 0
    while True:
        try:
            # Gunakan timeout agar tidak mudah 'database is locked'
            conn = sqlite3.connect(DB_NAME, timeout=10, detect_types=sqlite3.PARSE_DECLTYPES)
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute("SELECT * FROM logs WHERE is_synced=0")
            logs = cur.fetchall()
            
            for log in logs:
                ref = db.reference('absensi_log')
                waktu_masuk_obj = datetime.datetime.strptime(log['waktu_masuk'], DB_TIME_FORMAT) if log['waktu_masuk'] else None
                waktu_keluar_obj = datetime.datetime.strptime(log['waktu_keluar'], DB_TIME_FORMAT) if log['waktu_keluar'] else None
                row_id = log['id']
                nama_user = log['nama']
                status = log['status']
                
                # --- GUARD: Skip log korup tanpa waktu_masuk agar tidak crash loop ---
                if not waktu_masuk_obj:
                    print(f" -> [SKIP] Log ID={row_id} ({nama_user}) tidak punya waktu_masuk. Ditandai synced.")
                    cur.execute("UPDATE logs SET is_synced=1 WHERE id=?", (row_id,))
                    conn.commit()
                    continue
                
                current_ts = int(datetime.datetime.now().timestamp() * 1000)
                if current_ts <= last_used_timestamp:
                    current_ts = last_used_timestamp + 1
                last_used_timestamp = current_ts
                now_timestamp_ms = current_ts

                if status.upper() == "MASUK":
                    existing_fb_key = log['firebase_key']
                    
                    data = {
                        "nama_asisten": nama_user, "id_asisten_kampus": log['id_asisten_kampus'],
                        "lokasi_lab": log['lokasi_lab'], "kelas": log['kelas'], "status": "Masuk",
                        "tanggal": waktu_masuk_obj.strftime("%Y-%m-%d"),
                        "time_in": waktu_masuk_obj.strftime("%H:%M:%S"), "time_out": "",
                        "timestamp_ms": now_timestamp_ms
                    }
                    
                    if existing_fb_key:
                        # --- KASUS EDIT: Log sudah pernah di-push, UPDATE data di key yang sama ---
                        ref.child(existing_fb_key).update(data)
                        cur.execute("UPDATE logs SET is_synced=1 WHERE id=?", (row_id,))
                        conn.commit()
                        print(f" -> [UPDATE-MASUK] Updated: {nama_user} (MASUK DIEDIT) - Key: {existing_fb_key}")
                    else:
                        # --- KASUS BARU: Log belum pernah di-push, buat Key baru ---
                        new_ref = ref.push() # 1. Generate Key Firebase tanpa mengirim data
                        fb_key = new_ref.key
                        
                        # 2. Simpan dan COMMIT Key tersebut di SQLite lokal SEKARANG JUGA
                        cur.execute("UPDATE logs SET firebase_key=?, is_synced=1 WHERE id=?", (fb_key, row_id))
                        conn.commit() 
                        
                        # 3. SETELAH lokal aman, baru tembak datanya ke Firebase
                        new_ref.set(data)
                        print(f" -> [PUSH] Uploaded: {nama_user} (MASUK) - Key: {fb_key}")
                    
                elif status.startswith("Keluar"):
                    firebase_key = log['firebase_key']
                    
                    if firebase_key:
                        # --- PERBAIKAN: Kirim SEMUA field agar edit apapun ter-sync ---
                        update_data = {
                            "nama_asisten": nama_user,
                            "id_asisten_kampus": log['id_asisten_kampus'],
                            "lokasi_lab": log['lokasi_lab'],
                            "kelas": log['kelas'],
                            "tanggal": waktu_masuk_obj.strftime("%Y-%m-%d"),
                            "time_in": waktu_masuk_obj.strftime("%H:%M:%S"),
                            "time_out": waktu_keluar_obj.strftime("%H:%M:%S"),
                            "status": status,
                            "timestamp_ms": now_timestamp_ms 
                        }
                        # Update Cloud dulu, baru lokal
                        ref.child(firebase_key).update(update_data)
                        cur.execute("UPDATE logs SET is_synced=1 WHERE id=?", (row_id,))
                        conn.commit() # Commit instan
                        print(f" -> [UPDATE] Uploaded: {nama_user} ({status})")
                    else:
                        # --- KASUS SPESIAL OFFLINE (KUNCI LOKAL DULU) ---
                        data = {
                            "nama_asisten": nama_user, "id_asisten_kampus": log['id_asisten_kampus'],
                            "lokasi_lab": log['lokasi_lab'], "kelas": log['kelas'], "status": status,
                            "tanggal": waktu_masuk_obj.strftime("%Y-%m-%d"),
                            "time_in": waktu_masuk_obj.strftime("%H:%M:%S"),
                            "time_out": waktu_keluar_obj.strftime("%H:%M:%S"),
                            "timestamp_ms": now_timestamp_ms
                        }
                        new_ref = ref.push() # 1. Generate Key
                        fb_key = new_ref.key
                        
                        # 2. Simpan dan COMMIT
                        cur.execute("UPDATE logs SET firebase_key=?, is_synced=1 WHERE id=?", (fb_key, row_id))
                        conn.commit()
                        
                        # 3. Kirim ke Firebase
                        new_ref.set(data)
                        print(f" -> [PUSH-OFFLINE] Sesi offline penuh untuk {nama_user}. Uploading...")

            conn.close()
        except Exception as e:
            print(f"[UPLOAD ERROR] {e}")
            try: conn.close() 
            except: pass
        time.sleep(5)

def task_upload_master_data():
    """Mengunggah antrean perubahan Master Data (Asisten & Jadwal) ke Firebase"""
    print("[UPLOAD-MASTER] Thread Upload Master Data Berjalan...")
    while True:
        try:
            conn = sqlite3.connect(DB_NAME, timeout=10)
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            
            # --- 1. UPLOAD ASISTEN ---
            cur.execute("SELECT * FROM users WHERE is_synced=0")
            unsynced_users = cur.fetchall()
            for user in unsynced_users:
                u_dict = dict(user)
                id_kampus = u_dict['id_asisten_kampus']
                action = u_dict['sync_action']
                
                ref = db.reference(f'asisten_master/{id_kampus}') # Memaksa Firebase pakai ID Kampus sebagai Key
                
                if action in ['INSERT', 'UPDATE']:
                    data = {
                        'fingerprint_id': u_dict['fingerprint_id'],
                        'nama': u_dict['nama'],
                        'id_asisten_kampus': id_kampus,
                        'hak_akses': u_dict['hak_akses']
                    }
                    ref.set(data) # Gunakan set() untuk memaksakan key/nama folder
                    cur.execute("UPDATE users SET is_synced=1, sync_action=NULL WHERE fingerprint_id=?", (u_dict['fingerprint_id'],))
                    print(f" ->[MASTER-PUSH] Asisten {u_dict['nama']} diupload ke Firebase.")
                    
                elif action == 'DELETE':
                    ref.delete()
                    cur.execute("DELETE FROM users WHERE fingerprint_id=?", (u_dict['fingerprint_id'],))
                    print(f" -> [MASTER-DEL] Asisten {u_dict['nama']} dihapus dari Firebase & Lokal.")
                    
            # --- 2. UPLOAD JADWAL ---
            cur.execute("SELECT * FROM schedules WHERE is_synced=0")
            unsynced_schedules = cur.fetchall()
            for sched in unsynced_schedules:
                s_dict = dict(sched)
                id_j = s_dict['id_jadwal']
                action = s_dict['sync_action']
                
                ref = db.reference(f'jadwal_kelas/{id_j}') # Memaksa Firebase pakai ID Jadwal sebagai Key
                
                if action in ['INSERT', 'UPDATE']:
                    data = {
                        'hari': s_dict['hari'],
                        'is_online': bool(s_dict['is_online']),
                        'jam_mulai': s_dict['jam_mulai'],
                        'jam_selesai': s_dict['jam_selesai'],
                        'lokasi_lab': s_dict['lokasi_lab'],
                        'nama_kelas': s_dict['nama_kelas']
                    }
                    ref.set(data)
                    cur.execute("UPDATE schedules SET is_synced=1, sync_action=NULL WHERE id_jadwal=?", (id_j,))
                    print(f" -> [MASTER-PUSH] Jadwal {s_dict['nama_kelas']} diupload ke Firebase.")
                    
                elif action == 'DELETE':
                    ref.delete()
                    cur.execute("DELETE FROM schedules WHERE id_jadwal=?", (id_j,))
                    print(f" -> [MASTER-DEL] Jadwal {s_dict['nama_kelas']} dihapus dari Firebase & Lokal.")

            conn.commit(); conn.close()
        except Exception as e:
            print(f"[UPLOAD-MASTER ERROR] {e}")
        
        time.sleep(5) # Cek antrean setiap 5 detik

# --- FUNGSI BARU UNTUK LISTENER ---
def stream_device_control_listener():
    """
    Mendengarkan perubahan pada node 'device_control' di Firebase secara real-time.
    Dilengkapi dengan fitur Self-Healing agar selalu siaga menerima perintah Admin.
    """
    def listener(event):
        global CACHED_DEVICE_CONTROL
        # event.path akan berisi '/' jika seluruh node berubah, atau '/LAB_AP' jika hanya satu lab yang berubah
        # event.data akan berisi data yang baru
        print(f"[CONTROL-STREAM] Perubahan Terdeteksi pada Device Control: Path={event.path}")

        # Logika ini sudah bagus, tetapi perlu disempurnakan sedikit untuk menangani semua kasus
        if event.path == "/":
             # Jika seluruh node device_control di-replace atau saat pertama kali dijalankan
             if event.data:
                CACHED_DEVICE_CONTROL = event.data
             else: # Jika node dihapus
                CACHED_DEVICE_CONTROL = {}
        else:
             # Jika hanya sub-node (misal /LAB_AP/mode atau /LAB_AP) yang berubah
             path_parts = event.path.strip("/").split("/")
             lab_name = path_parts[0] # Ambil bagian pertama sebagai nama lab

             # Kita perlu membangun kembali state lab tersebut dari data yang diterima
             # event.data hanya berisi nilai yang berubah, bukan seluruh objek lab
             # Jadi, cara terbaik adalah mengambil ulang seluruh data untuk lab tersebut
             if lab_name:
                 updated_lab_data = db.reference(f'device_control/{lab_name}').get()
                 CACHED_DEVICE_CONTROL[lab_name] = updated_lab_data

        print(f"[CONTROL-STREAM] Cache Device Control diperbarui: {CACHED_DEVICE_CONTROL}")

    # Setup listener
    while True:
        print("[CONTROL-STREAM] Memulai/Me-restart koneksi Listener Device Control...")
        try:
            my_stream = firebase_admin.db.reference('device_control').listen(listener)
            
            # Tahan thread di sini agar terus mendengarkan
            while True:
                time.sleep(60)
                
        except Exception as e:
            print(f"[CONTROL-STREAM PUTUS/ERROR] {e}. Me-restart Listener Device Control dalam 10 detik...")
            try: my_stream.close()
            except: pass
            time.sleep(10)

# ================= API SCAN (REVISI DENGAN MODE DELETE) =================
@app.route('/api/scan', methods=['GET'])
def scan_jari():
    fp_id = request.args.get('id', type=int)
    lab_name = request.args.get('lab', 'UNKNOWN')
    
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    
    lab_control = CACHED_DEVICE_CONTROL.get(lab_name, {})

    # --- BAGIAN 1: CEK PERINTAH KHUSUS (ENROLL & DELETE) ---
    current_mode = lab_control.get('mode', 'absensi')
    target_id = lab_control.get('target_id', 0) 

    if current_mode == 'enroll' and target_id > 0:
        conn.close()
        print(f"[CONTROL] Mengirim perintah ENROLL ID {target_id} ke {lab_name}.")
        db.reference(f'device_control/{lab_name}').update({'mode': 'absensi', 'target_id': 0})
        return f"ENROLL|{target_id}"
    
    elif current_mode == 'delete' and target_id > 0:
        conn.close()
        print(f"[CONTROL] Mengirim perintah DELETE ID {target_id} ke {lab_name}.")
        db.reference(f'device_control/{lab_name}').update({'mode': 'absensi', 'target_id': 0})
        return f"DELETE|{target_id}"

    # --- BAGIAN 2: LOGIKA ABSENSI NORMAL ---
    
    cur.execute("SELECT nama, id_asisten_kampus, hak_akses FROM users WHERE fingerprint_id=?", (fp_id,))
    user = cur.fetchone()
    
    if not user:
        conn.close()
        if lab_control.get('mode') == 'enroll':
             enroll_id_target = lab_control.get('target_id', 0)
             db.reference(f'device_control/{lab_name}').update({'mode': 'absensi', 'target_id': 0})
             return f"ENROLL|{enroll_id_target}"
        return "GAGAL|ID Tidak Terdaftar"
    
    nama = user[0]
    id_kampus = user[1]
    hak_akses_str = user[2]

    list_hak_akses = [h.strip() for h in hak_akses_str.split(',')]
    if lab_name not in list_hak_akses:
        conn.close()
        print(f"[REJECT] {nama} mencoba absen di {lab_name} tapi tidak punya hak akses.")
        return f"GAGAL|Akses Ditolak"
    # Cek Active Session (di lab manapun)
    cur.execute("SELECT log_db_id, jam_selesai_kelas, lokasi_lab FROM active_sessions WHERE fingerprint_id=?", (fp_id,))
    active_session = cur.fetchone()
    
    now = datetime.datetime.now()
    
    if active_session:
        # --- PROSES TAP OUT atau PENOLAKAN ---
        log_db_id, jam_selesai_str, lab_aktif = active_session
        
        if lab_name == lab_aktif:
            # Tap Out Sah
            waktu_selesai_kelas = datetime.datetime.strptime(jam_selesai_str, "%Y-%m-%d %H:%M:%S")
            waktu_keluar_final = min(now, waktu_selesai_kelas)
            
            cur.execute("UPDATE logs SET status='Keluar', waktu_keluar=?, is_synced=0 WHERE id=?", (waktu_keluar_final.strftime(DB_TIME_FORMAT), log_db_id))
            cur.execute("DELETE FROM active_sessions WHERE fingerprint_id=?", (fp_id,))
            conn.commit(); conn.close()
            return f"KELUAR|{nama}"
        else:
            # Tap In Ilegal
            conn.close()
            return f"GAGAL|Aktif di {lab_aktif}"
        
    else:
        # --- PROSES TAP IN ---
        jadwal, waktu_mulai_kelas, waktu_selesai_kelas = get_current_schedule(lab_name)
        if not jadwal:
            conn.close()
            return "GAGAL|Tidak Ada Jadwal"
            
        nama_kelas = jadwal['nama_kelas']
        waktu_masuk_final = max(now, waktu_mulai_kelas)

        cur.execute('''INSERT INTO logs 
            (fingerprint_id, nama, id_asisten_kampus, waktu_masuk, status, lokasi_lab, kelas, is_synced)
            VALUES (?, ?, ?, ?, ?, ?, ?, 0)''', 
            (fp_id, nama, id_kampus, waktu_masuk_final.strftime(DB_TIME_FORMAT), 'MASUK', lab_name, nama_kelas))
        
        last_id = cur.lastrowid
        
        cur.execute('''INSERT INTO active_sessions 
            (fingerprint_id, nama, id_asisten_kampus, waktu_masuk, jam_selesai_kelas, lokasi_lab, kelas, log_db_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
            (fp_id, nama, id_kampus, 
            waktu_masuk_final.strftime(DB_TIME_FORMAT),  # Format ini
            waktu_selesai_kelas.strftime(DB_TIME_FORMAT), # dan format ini
            lab_name, nama_kelas, last_id))
            
        conn.commit(); conn.close()
        return f"MASUK|{nama}"

# --- DASHBOARD UTAMA (AT A GLANCE) ---
@app.route('/home')
@app.route('/')
@app.route('/home/<month_filter>')
def dashboard_home(month_filter=None):
    # Ambil Data Dasar untuk Navigasi (Ringan)
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    global CACHED_JADWAL 
    raw_labs = [jadwal.get('lokasi_lab') for jadwal in CACHED_JADWAL if jadwal.get('lokasi_lab')]
    all_labs_for_nav = sorted(list(set([str(lab).strip() for lab in raw_labs])))
    conn.close()

    # Hitung Navigasi Bulan
    if not month_filter: target_date = datetime.date.today()
    else: 
        try: target_date = datetime.datetime.strptime(month_filter, '%Y-%m').date()
        except: target_date = datetime.date.today()
    
    current_month = target_date.strftime('%Y-%m')
    prev_month = (target_date - relativedelta(months=1)).strftime('%Y-%m')
    next_month = (target_date + relativedelta(months=1)).strftime('%Y-%m')

    # [HTML REMOVED - Moved to templates/]
    return render_template('dashboard_home.html', 
                           current_month=current_month, 
                           prev_month=prev_month, 
                           next_month=next_month)

# ==============================================================================
# ROUTE LOG HARIAN (VERSI ADMIN CRUD DENGAN DROPDOWN MASTER DATA)
# ==============================================================================
@app.route('/log')
@app.route('/log/<month_filter>')
def dashboard_log(month_filter=None):
    is_admin = session.get('is_admin', False)

    if not month_filter: target_date = datetime.date.today()
    else: 
        try: target_date = datetime.datetime.strptime(month_filter, '%Y-%m').date()
        except: target_date = datetime.date.today()
            
    current_month_str = target_date.strftime('%Y-%m')
    prev_month_str = (target_date - relativedelta(months=1)).strftime('%Y-%m')
    next_month_str = (target_date + relativedelta(months=1)).strftime('%Y-%m')

    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    
    global CACHED_JADWAL 
    raw_labs =[jadwal.get('lokasi_lab') for jadwal in CACHED_JADWAL if jadwal.get('lokasi_lab')]
    all_labs_for_nav = sorted(list(set([str(lab).strip() for lab in raw_labs])))
    
    # --- AMBIL MASTER DATA UNTUK DROPDOWN FORM ---
    master_asisten =[]
    master_jadwal =[]
    
    if is_admin:
        # Ambil daftar Asisten aktif
        cur.execute("SELECT id_asisten_kampus, nama FROM users WHERE sync_action != 'DELETE' OR sync_action IS NULL ORDER BY fingerprint_id ASC")
        master_asisten =[dict(row) for row in cur.fetchall()]
        
        # Ambil daftar Lab & Kelas unik
        cur.execute("SELECT DISTINCT lokasi_lab, nama_kelas FROM schedules WHERE sync_action != 'DELETE' OR sync_action IS NULL ORDER BY lokasi_lab ASC, nama_kelas ASC")
        master_jadwal = [dict(row) for row in cur.fetchall()]
        
    conn.close()

    # [HTML REMOVED - Moved to templates/]
    return render_template('log_harian.html', 
                           is_admin=is_admin,
                           master_asisten=master_asisten,
                           master_jadwal=master_jadwal,
                           current_month=current_month_str, 
                           prev_month=prev_month_str, 
                           next_month=next_month_str)

# --- DASHBOARD ROUTE: REKAP GAJI PER LAB ---
@app.route('/rekap/<lab_name>')
@app.route('/rekap/<lab_name>/<month_filter>')
def dashboard_rekap(lab_name, month_filter=None):
    # Setup Navigasi & Bulan
    if not month_filter: target_date = datetime.date.today()
    else: 
        try: target_date = datetime.datetime.strptime(month_filter, '%Y-%m').date()
        except: target_date = datetime.date.today()

    current_month_str = target_date.strftime('%Y-%m')
    prev_month_str = (target_date - relativedelta(months=1)).strftime('%Y-%m')
    next_month_str = (target_date + relativedelta(months=1)).strftime('%Y-%m')
    
    # Ambil List Lab untuk Navbar
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    global CACHED_JADWAL 
    raw_labs = [jadwal.get('lokasi_lab') for jadwal in CACHED_JADWAL if jadwal.get('lokasi_lab')]
    all_labs_for_nav = sorted(list(set([str(lab).strip() for lab in raw_labs])))
    conn.close()

    # [HTML REMOVED - Moved to templates/]
    return render_template('rekap_gaji.html', 
                           lab_name=lab_name,
                           lab_name_display=lab_name.replace('_', ' '),
                           current_month=current_month_str, 
                           prev_month=prev_month_str, 
                           next_month=next_month_str)

# --- DASHBOARD ROUTE: - SESI AKTIF ---
@app.route('/active')
@app.route('/active/<lab_filter>')
def dashboard_active(lab_filter=None):
    # Setup Navigasi (Python) - Tetap diperlukan untuk render awal
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    global CACHED_JADWAL 
    raw_labs = [jadwal.get('lokasi_lab') for jadwal in CACHED_JADWAL if jadwal.get('lokasi_lab')]
    all_labs_for_nav = sorted(list(set([str(lab).strip() for lab in raw_labs])))
    
    raw_active_labs = [jadwal.get('lokasi_lab') for jadwal in CACHED_JADWAL if jadwal.get('lokasi_lab')]
    all_labs = sorted(list(set([str(lab).strip() for lab in raw_active_labs])))
    conn.close()
    
    # Kita tidak perlu query sesi aktif di sini lagi, karena akan diambil via JS
    
    # [HTML REMOVED - Moved to templates/]
    return render_template('active_sessions.html', all_labs=all_labs, lab_filter=lab_filter)

# --- DASHBOARD ROUTE: PROFIL ASISTEN SPESIFIK BULAN ---
@app.route('/profil/<id_asisten_kampus>/<lab_name>/<month_filter>')
def profil_asisten(id_asisten_kampus, lab_name, month_filter):
    # --- LOGIKA TANGGAL ---
    # Kita hanya butuh bulan saat ini, tidak perlu hitung prev/next lagi
    try:
        target_date = datetime.datetime.strptime(month_filter, '%Y-%m').date()
    except ValueError:
        target_date = datetime.date.today()
    
    current_month_str = target_date.strftime('%Y-%m')

    # --- AMBIL NAMA & LIST LAB UNTUK JUDUL/NAVBAR ---
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    # Ambil nama asisten
    cur.execute("SELECT nama FROM users WHERE id_asisten_kampus = ?", (id_asisten_kampus,))
    row = cur.fetchone()
    nama_asisten = row[0] if row else id_asisten_kampus
    
    # Ambil list lab untuk navbar utama
    global CACHED_JADWAL 
    raw_labs = [jadwal.get('lokasi_lab') for jadwal in CACHED_JADWAL if jadwal.get('lokasi_lab')]
    all_labs_for_nav = sorted(list(set([str(lab).strip() for lab in raw_labs])))
    conn.close()

    # Format Judul Halaman
    bulan_map = ["", "Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
    bulan_nama = bulan_map[target_date.month]
    tahun = target_date.year
    judul_halaman = f"Riwayat Absensi - {nama_asisten} di {lab_name.replace('_', ' ')} - Bulan {bulan_nama} {tahun}"

    # --- TEMPLATE HTML ---
    # [HTML REMOVED - Moved to templates/]
    return render_template('profil_asisten.html', 
                           judul_halaman=judul_halaman, 
                           id_asisten_kampus=id_asisten_kampus, 
                           lab_name=lab_name,
                           month_filter=current_month_str)


# --- API ROUTE: FORCE CHECKOUT ---
@app.route('/force_checkout/<int:fp_id>')
def force_checkout(fp_id):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    lab_filter_sekarang = request.args.get('filter', 'semua')
    # Cari sesi aktif berdasarkan fingerprint_id
    cur.execute("SELECT log_db_id FROM active_sessions WHERE fingerprint_id=?", (fp_id,))
    active = cur.fetchone()
    
    if active:
        log_db_id = active[0]
        now = datetime.datetime.now()
        
        # Update log dengan status baru
        cur.execute("UPDATE logs SET status='Keluar (Forced)', waktu_keluar=?, is_synced=0 WHERE id=?", 
                   (now.strftime(DB_TIME_FORMAT), log_db_id))
        
        # Hapus dari sesi aktif
        cur.execute("DELETE FROM active_sessions WHERE fingerprint_id=?", (fp_id,))
        
        conn.commit()
        print(f"[ADMIN] Berhasil Force Tap-Out untuk ID: {fp_id}")
    else:
        print(f"[ADMIN] Gagal Force Tap-Out, sesi untuk ID {fp_id} tidak ditemukan.")
        
    conn.close()
    
    if lab_filter_sekarang and lab_filter_sekarang != 'semua':
        # Jika ada filter lab, kembalikan ke /active/NAMA_LAB
        return redirect(url_for('dashboard_active', lab_filter=lab_filter_sekarang))
    else:
        # Jika filternya 'semua' atau kosong, kembalikan ke default /active
        return redirect(url_for('dashboard_active'))

@app.route('/export/log/<month_filter>')
def export_log(month_filter):
    try:
        # 1. AMBIL DATA
        conn = sqlite3.connect(DB_NAME)
        query = "SELECT * FROM logs WHERE strftime('%Y-%m', waktu_masuk) = ? ORDER BY id DESC"
        df = pd.read_sql_query(query, conn, params=(month_filter,))
        conn.close()

        if df.empty:
            return "Tidak ada data untuk diexport", 404

        # 2. PROSES DATA 
        tarif_per_menit = 281.25

        def process_row(row):
            # A. Format Tanggal, Time IN, Time OUT
            try:
                dt_masuk = datetime.datetime.strptime(row['waktu_masuk'], DB_TIME_FORMAT)
                tanggal = dt_masuk.strftime('%Y-%m-%d')
                time_in = dt_masuk.strftime('%H:%M:%S')
            except (ValueError, TypeError):
                tanggal, time_in = "", ""

            try:
                dt_keluar = datetime.datetime.strptime(row['waktu_keluar'], DB_TIME_FORMAT)
                time_out = dt_keluar.strftime('%H:%M:%S')
            except (ValueError, TypeError):
                time_out = ""

            # B. Hitung Durasi & Gaji
            menit = hitung_durasi_menit(row['waktu_masuk'], row['waktu_keluar'])
            
            if menit > 0:
                jam_dur = int(menit // 60)
                menit_dur = int(menit % 60)
                durasi_str = f"{jam_dur:02d}:{menit_dur:02d}"
                gaji = menit * tarif_per_menit
            else:
                durasi_str = "00:00"
                gaji = 0

            return pd.Series([tanggal, time_in, time_out, durasi_str, gaji])

        df[['Tanggal', 'Time IN', 'Time OUT', 'Durasi (HH:MM)', 'Gaji']] = df.apply(process_row, axis=1)

        # 3. RAPIKAN KOLOM
        cols_to_keep = ['id', 'nama', 'id_asisten_kampus', 'kelas', 'lokasi_lab', 
                        'Tanggal', 'Time IN', 'Time OUT', 'Durasi (HH:MM)', 'Gaji', 'status']
        df = df[cols_to_keep]

        df = df.rename(columns={
            'id': 'No. Log', 'nama': 'Nama', 'id_asisten_kampus': 'ID Asisten',
            'kelas': 'Kelas', 'lokasi_lab': 'Lokasi Lab', 'status': 'Status'
        })

        # 4. EXPORT DENGAN STYLING (OPENPYXL)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Log Harian')
            
            workbook = writer.book
            worksheet = writer.sheets['Log Harian']
            
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            header_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            center_align = Alignment(horizontal='center', vertical='center')
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))

            # A. Styling Header
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

            # B. Styling Data
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.border = thin_border
                    if cell.column_letter not in ['B']: 
                        cell.alignment = center_align
                    
                    # Format Rupiah Kolom Gaji (J)
                    if cell.column == 10: 
                        cell.number_format = '"Rp"#,##0.00'

            # C. Auto-adjust Column Width (DIPERBAIKI)
            for column in worksheet.columns:
                max_length = 0
                column_name = column[0].column_letter 
                
                # Hitung panjang teks terpanjang di kolom ini
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # --- PERBAIKAN DI SINI ---
                # Khusus Kolom Gaji (J), kita paksa lebarnya agar muat format Rp
                if column_name == 'J':
                    adjusted_width = 25 # Lebar manual yang cukup luas
                else:
                    adjusted_width = (max_length + 3) # Tambah buffer dikit biar gak mepet
                
                worksheet.column_dimensions[column_name].width = adjusted_width

        output.seek(0)

        return send_file(output, as_attachment=True, 
                         download_name=f"Log_Absensi_{month_filter}.xlsx", 
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
    except Exception as e:
        return f"Error export log: {str(e)}"

# 2. EXPORT REKAP GAJI
@app.route('/export/rekap/<lab_name>/<month_filter>')
def export_rekap(lab_name, month_filter):
    try:
        # 1. AMBIL DATA
        conn = sqlite3.connect(DB_NAME)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        query = "SELECT * FROM logs WHERE lokasi_lab = ? AND strftime('%Y-%m', waktu_masuk) = ?"
        cur.execute(query, (lab_name, month_filter))
        logs_lab = cur.fetchall()
        conn.close()
        
        if not logs_lab:
            return "Tidak ada data rekap untuk bulan ini", 404

        # 2. PROSES KALKULASI (Sama seperti Dashboard)
        rekap = {}
        for log in logs_lab:
            nama = log['nama']
            if nama not in rekap:
                rekap[nama] = {'id_asisten': log['id_asisten_kampus'], 'total_menit': 0}
            
            menit = hitung_durasi_menit(log['waktu_masuk'], log['waktu_keluar'])
            rekap[nama]['total_menit'] += menit

        tarif_per_menit = 281.25
        anggaran_maks = 1080000
        total_kelebihan_gaji = 0
        total_sisa_anggaran = 0
        
        temp_data = []
        
        # Hitung Dasar
        for nama, data in rekap.items():
            total_gaji = data['total_menit'] * tarif_per_menit
            gaji_pokok = min(total_gaji, anggaran_maks)
            sisa_anggaran = max(0, anggaran_maks - total_gaji)
            kelebihan_gaji = max(0, total_gaji - anggaran_maks)
            
            total_sisa_anggaran += sisa_anggaran
            total_kelebihan_gaji += kelebihan_gaji
            
            temp_data.append({
                'Nama': nama, 
                'ID Asisten': data['id_asisten'],
                'Total Gaji': total_gaji,
                'Gaji Pokok': gaji_pokok,
                'Sisa Anggaran': sisa_anggaran,
                'Kelebihan Gaji': kelebihan_gaji,
                # Placeholder
                'Proporsi Bonus': 0, 
                'Bonus Diterima': 0, 
                'GAJI FINAL': 0
            })

        # Hitung Bonus & Final
        for item in temp_data:
            proporsi = (item['Kelebihan Gaji'] / total_kelebihan_gaji) if total_kelebihan_gaji > 0 else 0
            bonus_diterima = proporsi * total_sisa_anggaran
            
            item['Proporsi Bonus'] = proporsi
            item['Bonus Diterima'] = bonus_diterima
            item['GAJI FINAL'] = item['Gaji Pokok'] + bonus_diterima

        # 3. URUTKAN & BERI NOMOR (Agar sama persis dengan Dashboard)
        # Urutkan berdasarkan Gaji Final Tertinggi
        temp_data_sorted = sorted(temp_data, key=lambda x: x['GAJI FINAL'], reverse=True)
        
        # Tambahkan kolom "No."
        final_data_list = []
        for index, item in enumerate(temp_data_sorted):
            # Buat dictionary baru dengan urutan kunci yang benar
            row = {
                'No.': index + 1,
                'Nama': item['Nama'],
                'ID Asisten': item['ID Asisten'],
                'Total Gaji': item['Total Gaji'],
                'Gaji Pokok': item['Gaji Pokok'],
                'Sisa Anggaran': item['Sisa Anggaran'],
                'Kelebihan Gaji': item['Kelebihan Gaji'],
                'Proporsi Bonus': item['Proporsi Bonus'],
                'Bonus Diterima': item['Bonus Diterima'],
                'GAJI FINAL': item['GAJI FINAL']
            }
            final_data_list.append(row)

        # Buat DataFrame
        df = pd.DataFrame(final_data_list)

        # 4. EXPORT DENGAN STYLING (OPENPYXL)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            sheet_name = f"Rekap {lab_name}"
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            # Definisi Style
            header_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            center_align = Alignment(horizontal='center', vertical='center')
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))

            # A. Styling Header
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

            # B. Styling Data Row
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.border = thin_border
                    
                    # Kolom Nama (B) rata kiri, sisanya rata tengah
                    if cell.column_letter == 'B':
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    else:
                        cell.alignment = center_align
                    
                    # Format Rupiah (Kolom D, E, F, G, I, J)
                    # D=4, E=5, F=6, G=7, I=9, J=10
                    if cell.column in [4, 5, 6, 7, 9, 10]:
                        cell.number_format = '"Rp"#,##0.00'
                    
                    # Format Persen (Kolom H / ke-8)
                    if cell.column == 8:
                        cell.number_format = '0.00%'
                    
                    # Bold untuk Gaji Final (Kolom J)
                    if cell.column == 10:
                        cell.font = Font(bold=True)

            # C. Auto-adjust Column Width
            for column in worksheet.columns:
                max_length = 0
                column_name = column[0].column_letter 
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Manual override untuk kolom Uang agar lebar pas
                if column_name in ['D', 'E', 'F', 'G', 'I', 'J']:
                    adjusted_width = 20
                else:
                    adjusted_width = (max_length + 2)
                
                worksheet.column_dimensions[column_name].width = adjusted_width

        output.seek(0)

        return send_file(output, as_attachment=True, 
                         download_name=f"Rekap_Gaji_{lab_name}_{month_filter}.xlsx", 
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        return f"Error export rekap: {str(e)}"

# 3. EXPORT PROFIL ASISTEN
@app.route('/export/profil/<id_asisten>/<lab_name>/<month_filter>')
def export_profil(id_asisten, lab_name, month_filter):
    try:
        # 1. AMBIL DATA (Sama seperti tampilan web)
        conn = sqlite3.connect(DB_NAME)
        query = """
            SELECT * FROM logs 
            WHERE id_asisten_kampus = ? AND lokasi_lab = ? AND strftime('%Y-%m', waktu_masuk) = ? 
            ORDER BY id DESC
        """
        df = pd.read_sql_query(query, conn, params=(id_asisten, lab_name, month_filter))
        conn.close()

        if df.empty:
            return "Tidak ada data riwayat untuk diexport", 404

        # 2. PROSES DATA (Agar kolom sama persis dengan Web)
        tarif_per_menit = 281.25
        
        def process_row(row):
            # Format Tanggal & Jam
            try:
                dt_masuk = datetime.datetime.strptime(row['waktu_masuk'], DB_TIME_FORMAT)
                tanggal = dt_masuk.strftime('%Y-%m-%d')
                time_in = dt_masuk.strftime('%H:%M:%S')
            except (ValueError, TypeError):
                tanggal, time_in = "", ""

            try:
                dt_keluar = datetime.datetime.strptime(row['waktu_keluar'], DB_TIME_FORMAT)
                time_out = dt_keluar.strftime('%H:%M:%S')
            except (ValueError, TypeError):
                time_out = ""

            # Hitung Durasi & Gaji
            menit = hitung_durasi_menit(row['waktu_masuk'], row['waktu_keluar'])
            
            if menit > 0:
                jam_dur = int(menit // 60)
                menit_dur = int(menit % 60)
                durasi_str = f"{jam_dur:02d}:{menit_dur:02d}"
                gaji = menit * tarif_per_menit
            else:
                durasi_str = "00:00"
                gaji = 0

            return pd.Series([tanggal, time_in, time_out, durasi_str, gaji])

        # Terapkan processing
        df[['Tanggal', 'Time IN', 'Time OUT', 'Durasi (HH:MM)', 'Gaji']] = df.apply(process_row, axis=1)

        # 3. RAPIKAN KOLOM (Pilih dan Rename)
        cols_to_keep = ['id', 'nama', 'id_asisten_kampus', 'kelas', 'lokasi_lab', 
                        'Tanggal', 'Time IN', 'Time OUT', 'Durasi (HH:MM)', 'Gaji', 'status']
        df = df[cols_to_keep]

        df = df.rename(columns={
            'id': 'No. Log', 
            'nama': 'Nama', 
            'id_asisten_kampus': 'ID Asisten', 
            'kelas': 'Kelas',
            'lokasi_lab': 'Lokasi Lab',
            'status': 'Status'
        })
        nama_asisten_file = df.iloc[0]['Nama'] 
        # 4. EXPORT DENGAN STYLING (OPENPYXL)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            sheet_name = 'Riwayat Absensi'
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            header_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            center_align = Alignment(horizontal='center', vertical='center')
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))

            # A. Styling Header
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

            # B. Styling Data Row
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.border = thin_border
                    
                    if cell.column_letter != 'B': # Selain Nama, rata tengah
                        cell.alignment = center_align
                    
                    # Kolom Gaji (J / ke-10) -> Format Rupiah
                    if cell.column == 10: 
                        cell.number_format = '"Rp"#,##0.00'

            # C. Auto-adjust Column Width
            for column in worksheet.columns:
                max_length = 0
                column_name = column[0].column_letter 
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Paksa lebar kolom Gaji (J)
                if column_name == 'J':
                    adjusted_width = 25
                else:
                    adjusted_width = (max_length + 3)
                
                worksheet.column_dimensions[column_name].width = adjusted_width

        output.seek(0)

        # Nama file
        filename = f"Riwayat_{nama_asisten_file}_{lab_name}_{month_filter}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        return f"Error export profil: {str(e)}"

@app.route('/api/data/home')
@app.route('/api/data/home/<month_filter>')
def api_dashboard_data(month_filter=None):
    # 1. SETUP TANGGAL
    if not month_filter:
        target_date = datetime.date.today()
    else:
        try:
            target_date = datetime.datetime.strptime(month_filter, '%Y-%m').date()
        except ValueError:
            target_date = datetime.date.today()

    current_month_str = target_date.strftime('%Y-%m')

    # 2. QUERY DATABASE
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # Metrik Kartu
    cur.execute("SELECT COUNT(*) FROM active_sessions")
    sesi_aktif = cur.fetchone()[0]
    
    today_str = datetime.date.today().strftime('%Y-%m-%d')
    cur.execute("SELECT COUNT(*) FROM logs WHERE date(waktu_masuk) = ?", (today_str,))
    absensi_hari_ini = cur.fetchone()[0]
    
    cur.execute("SELECT lokasi_lab, COUNT(*) as total FROM logs WHERE date(waktu_masuk) = ? GROUP BY lokasi_lab ORDER BY total DESC LIMIT 1", (today_str,))
    lab_teramai_row = cur.fetchone()
    lab_teramai = lab_teramai_row['lokasi_lab'].replace('_', ' ') if lab_teramai_row else "N/A"

    # LOAD DATA PANDAS
    df = pd.read_sql_query("SELECT * FROM logs", conn)
    conn.close()

    # Default Kosong
    graph_pie = None
    graph_line = None
    graphs_bar = {}
    top_asisten = {}

    if not df.empty:
        # --- PEMBERSIHAN DATA (Logika Final) ---
        df['waktu_masuk_str'] = df['waktu_masuk'].astype(str).str.slice(0, 19)
        df['dt_masuk'] = pd.to_datetime(df['waktu_masuk_str'], format='%Y-%m-%d %H:%M:%S', errors='coerce')
        df['bulan_str'] = df['dt_masuk'].dt.strftime('%Y-%m')
        
        df_bulan = df[df['bulan_str'] == current_month_str].copy()
        
        # --- A. DATA HADIR (Pie & Line) ---
        df_hadir = df_bulan.dropna(subset=['dt_masuk']).copy()
        if not df_hadir.empty:
            df_hadir['lokasi_lab'] = df_hadir['lokasi_lab'].astype(str).str.strip()
            
            # 1. PIE CHART
            pie_counts = df_hadir['lokasi_lab'].value_counts().reset_index()
            pie_counts.columns = ['lokasi_lab', 'counts']
            
            graph_pie = {
                'data': [{
                    'labels': pie_counts['lokasi_lab'].tolist(),
                    'values': pie_counts['counts'].tolist(),
                    'type': 'pie',
                    'hole': 0.4
                }],
                'layout': {'title': 'Persentase Keramaian per Lab (Total Sesi)'}
            }

            # 2. LINE CHART
            start_date = target_date.replace(day=1)
            end_date = (start_date + relativedelta(months=1)) - relativedelta(days=1)
            all_days = [d.strftime('%Y-%m-%d') for d in pd.date_range(start=start_date, end=end_date, freq='D')]
            
            df_hadir['tanggal_str'] = df_hadir['dt_masuk'].dt.strftime('%Y-%m-%d')
            daily = df_hadir.groupby(['tanggal_str', 'lokasi_lab']).size().unstack(fill_value=0).reindex(all_days, fill_value=0)
            
            line_data = []
            all_labs = sorted(df_hadir['lokasi_lab'].unique().tolist())
            for lab in all_labs:
                if lab in daily.columns:
                    line_data.append({
                        'x': all_days,
                        'y': daily[lab].tolist(),
                        'mode': 'lines+markers',
                        'name': lab
                    })
            
            graph_line = {
                'data': line_data,
                'layout': {
                    'title': 'Tren Jumlah Absensi Harian',
                    'xaxis': {'type': 'category', 'tickmode': 'auto', 'nticks': 10, 'title': 'Tanggal'},
                    'yaxis': {'tickformat': 'd', 'title': 'Jumlah Sesi'}
                }
            }

        # --- B. DATA DURASI (Bar & Ranking) ---
        df_bulan['waktu_keluar_str'] = df_bulan['waktu_keluar'].astype(str).str.slice(0, 19)
        df_bulan['dt_keluar'] = pd.to_datetime(df_bulan['waktu_keluar_str'], format='%Y-%m-%d %H:%M:%S', errors='coerce')
        
        df_durasi = df_bulan.dropna(subset=['dt_masuk', 'dt_keluar']).copy()
        df_durasi = df_durasi[df_durasi['dt_keluar'] > df_durasi['dt_masuk']]
        
        if not df_durasi.empty:
            df_durasi['lokasi_lab'] = df_durasi['lokasi_lab'].astype(str).str.strip()
            # Hitung jam dari menit penuh
            df_durasi['durasi_jam'] = ((df_durasi['dt_keluar'] - df_durasi['dt_masuk']).dt.total_seconds() // 60) / 60.0
            
            agg = df_durasi.groupby(['lokasi_lab', 'nama'])['durasi_jam'].sum().reset_index()
            labs_durasi = sorted(agg['lokasi_lab'].unique().tolist())
            
            for lab in labs_durasi:
                lab_data = agg[agg['lokasi_lab'] == lab].sort_values(by='durasi_jam', ascending=False)
                
                if not lab_data.empty:
                    # Ranking Top 3
                    top_asisten[lab] = lab_data.head(3).to_dict('records')
                    
                    # Bar Chart Top 10
                    chart_data = lab_data.head(10)
                    names = chart_data['nama'].tolist()
                    hours = chart_data['durasi_jam'].tolist()
                    
                    graphs_bar[lab] = {
                        'data': [{
                            'x': names,
                            'y': hours,
                            'type': 'bar',
                            'text': [f'{x:.2f} Jam' for x in hours],
                            'textposition': 'auto',
                            'marker': {'color': '#3F51B5'}
                        }],
                        'layout': {
                            'title': f'Total Jam Kerja - {lab}',
                            'xaxis': {'title': 'Nama Asisten', 'categoryorder': 'array', 'categoryarray': names},
                            'yaxis': {'title': 'Total Jam (Desimal)'}
                        }
                    }

    return jsonify({
        'sesi_aktif': sesi_aktif,
        'absensi_hari_ini': absensi_hari_ini,
        'lab_teramai': lab_teramai,
        'graph_pie': graph_pie,
        'graph_line': graph_line,
        'graphs_bar': graphs_bar,
        'top_asisten': top_asisten
    })

# ==============================================================================
# API ENDPOINT: DATA LOG HARIAN (JSON)
# ==============================================================================
@app.route('/api/data/log')
@app.route('/api/data/log/<month_filter>')
def api_log_data(month_filter=None):
    # 1. SETUP TANGGAL
    if not month_filter:
        target_date = datetime.date.today()
    else:
        try:
            target_date = datetime.datetime.strptime(month_filter, '%Y-%m').date()
        except ValueError:
            target_date = datetime.date.today()
            
    current_month_str = target_date.strftime('%Y-%m')

    # 2. AMBIL DATA
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    
    # Query SQL dengan filter bulan
    query = "SELECT * FROM logs WHERE strftime('%Y-%m', waktu_masuk) = ? ORDER BY id DESC LIMIT 500"
    cur.execute(query, (current_month_str,))
    logs = cur.fetchall()
    conn.close()
    
    # 3. PROSES DATA (FORMATTING UNTUK JSON)
    processed_logs = []
    tarif_per_menit = 281.25
    
    for log in logs:
        log_dict = dict(log)
        
        # Format Tanggal & Jam Masuk
        if log_dict['waktu_masuk']:
            dt_masuk = datetime.datetime.strptime(log_dict['waktu_masuk'], DB_TIME_FORMAT)
            log_dict['tanggal_display'] = dt_masuk.strftime('%Y-%m-%d')
            log_dict['time_in_display'] = dt_masuk.strftime('%H:%M:%S')
        else:
            log_dict['tanggal_display'] = 'N/A'
            log_dict['time_in_display'] = ''

        # Format Jam Keluar
        if log_dict['waktu_keluar']:
            try:
                dt_keluar = datetime.datetime.strptime(log_dict['waktu_keluar'], DB_TIME_FORMAT)
                log_dict['time_out_display'] = dt_keluar.strftime('%H:%M:%S')
            except:
                log_dict['time_out_display'] = ''
        else:
            log_dict['time_out_display'] = ''
            
        # Hitung Durasi
        durasi_menit = hitung_durasi_menit(log_dict['waktu_masuk'], log_dict['waktu_keluar'])
        if durasi_menit > 0:
            jam = int(durasi_menit // 60)
            menit = int(durasi_menit % 60)
            log_dict['durasi_display'] = f"{jam:02d}:{menit:02d}"
        else:
            log_dict['durasi_display'] = "00:00"
            
        # Hitung Gaji
        gaji = 0
        if durasi_menit > 0:
            gaji = durasi_menit * tarif_per_menit
        
        # Kita format Rupiah di Python biar JS gak ribet
        log_dict['gaji_display'] = format_rupiah(gaji)
            
        processed_logs.append(log_dict)

    return jsonify({'logs': processed_logs})

@app.route('/api/data/rekap/<lab_name>')
@app.route('/api/data/rekap/<lab_name>/<month_filter>')
def api_rekap_data(lab_name, month_filter=None):
    # 1. SETUP TANGGAL
    if not month_filter:
        target_date = datetime.date.today()
    else:
        try:
            target_date = datetime.datetime.strptime(month_filter, '%Y-%m').date()
        except ValueError:
            target_date = datetime.date.today()
            
    current_month_str = target_date.strftime('%Y-%m')

    # 2. AMBIL DATA
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    
    # Query Data
    query = "SELECT * FROM logs WHERE lokasi_lab = ? AND strftime('%Y-%m', waktu_masuk) = ?"
    cur.execute(query, (lab_name, current_month_str))
    logs_lab = cur.fetchall()
    conn.close()
    
    # 3. PROSES KALKULASI GAJI (Sama Persis dengan Dashboard Lama)
    rekap = {}
    for log in logs_lab:
        nama = log['nama']
        if nama not in rekap:
            rekap[nama] = {'id_asisten': log['id_asisten_kampus'], 'total_menit': 0}
        
        menit = hitung_durasi_menit(log['waktu_masuk'], log['waktu_keluar'])
        rekap[nama]['total_menit'] += menit

    tarif_per_menit = 281.25
    anggaran_maks = 1080000
    total_sisa_anggaran = 0
    total_kelebihan_gaji = 0
    
    data_final = []
    
    # Kalkulasi Tahap 1
    for nama, data in rekap.items():
        total_gaji = data['total_menit'] * tarif_per_menit
        gaji_pokok = min(total_gaji, anggaran_maks)
        sisa_anggaran = max(0, anggaran_maks - total_gaji)
        kelebihan_gaji = max(0, total_gaji - anggaran_maks)
        
        total_sisa_anggaran += sisa_anggaran
        total_kelebihan_gaji += kelebihan_gaji
        
        data_final.append({
            'nama': nama, 
            'id_asisten': data['id_asisten'],
            'total_gaji_raw': total_gaji,      # Simpan raw value untuk sorting
            'gaji_pokok_raw': gaji_pokok,
            'sisa_anggaran_raw': sisa_anggaran,
            'kelebihan_gaji_raw': kelebihan_gaji,
            'total_gaji': format_rupiah(total_gaji),
            'gaji_pokok': format_rupiah(gaji_pokok),
            'sisa_anggaran': format_rupiah(sisa_anggaran),
            'kelebihan_gaji': format_rupiah(kelebihan_gaji)
        })

    # Kalkulasi Tahap 2 (Bonus)
    for data in data_final:
        proporsi = (data['kelebihan_gaji_raw'] / total_kelebihan_gaji) if total_kelebihan_gaji > 0 else 0
        bonus_diterima = proporsi * total_sisa_anggaran
        gaji_final = data['gaji_pokok_raw'] + bonus_diterima
        
        data['proporsi_bonus'] = f"{proporsi:.2%}"
        data['bonus_diterima'] = format_rupiah(bonus_diterima)
        data['gaji_final'] = format_rupiah(gaji_final)
        data['gaji_final_raw'] = gaji_final # Untuk sorting di JS jika perlu, atau sorting di sini

    # Urutkan berdasarkan Gaji Final Tertinggi
    data_final_sorted = sorted(data_final, key=lambda item: item['gaji_final_raw'], reverse=True)

    return jsonify({'data': data_final_sorted})

@app.route('/api/data/active')
@app.route('/api/data/active/<lab_filter>')
def api_active_sessions(lab_filter=None):
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    
    # Query SQL dengan filter (Sama seperti sebelumnya)
    if lab_filter and lab_filter != "semua":
        query = "SELECT * FROM active_sessions WHERE lokasi_lab = ? ORDER BY waktu_masuk DESC"
        cur.execute(query, (lab_filter,))
    else:
        query = "SELECT * FROM active_sessions ORDER BY waktu_masuk DESC"
        cur.execute(query)
        
    active_sessions = cur.fetchall()
    conn.close()
    
    # Proses data untuk JSON (termasuk kalkulasi durasi & pewarnaan)
    processed_sessions = []
    now = datetime.datetime.now()
    
    for sess in active_sessions:
        sess_dict = dict(sess)
        
        # Hitung Durasi Menit untuk Pewarnaan
        waktu_masuk_obj = datetime.datetime.strptime(sess_dict['waktu_masuk'], DB_TIME_FORMAT)
        durasi_menit = (now - waktu_masuk_obj).total_seconds() / 60
        
        # Tentukan Class CSS
        if durasi_menit > 90:
            sess_dict['row_class'] = 'status-merah'
        elif durasi_menit > 45:
            sess_dict['row_class'] = 'status-kuning'
        else:
            sess_dict['row_class'] = 'status-hijau'
            
        processed_sessions.append(sess_dict)

    return jsonify({'sessions': processed_sessions})

@app.route('/api/data/profil/<id_asisten_kampus>/<lab_name>/<month_filter>')
def api_profil_data(id_asisten_kampus, lab_name, month_filter):
    # 1. SETUP TANGGAL
    try:
        target_date = datetime.datetime.strptime(month_filter, '%Y-%m').date()
    except ValueError:
        target_date = datetime.date.today()
    
    current_month_str = target_date.strftime('%Y-%m')

    # 2. AMBIL DATA
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    
    query = """
        SELECT * FROM logs 
        WHERE id_asisten_kampus = ? AND lokasi_lab = ? AND strftime('%Y-%m', waktu_masuk) = ? 
        ORDER BY id DESC
    """
    cur.execute(query, (id_asisten_kampus, lab_name, current_month_str))
    logs = cur.fetchall()
    conn.close()

    # 3. PROSES DATA (FORMATTING)
    processed_logs = []
    tarif_per_menit = 281.25
    
    for log in logs:
        log_dict = dict(log)
        
        # Format Tanggal
        if log_dict['waktu_masuk']:
            dt_masuk = datetime.datetime.strptime(log_dict['waktu_masuk'], DB_TIME_FORMAT)
            log_dict['tanggal_display'] = dt_masuk.strftime('%Y-%m-%d')
            log_dict['time_in_display'] = dt_masuk.strftime('%H:%M:%S')
        else:
            log_dict['tanggal_display'] = 'N/A'
            log_dict['time_in_display'] = ''

        # Format Jam Keluar
        if log_dict['waktu_keluar']:
            try:
                dt_keluar = datetime.datetime.strptime(log_dict['waktu_keluar'], DB_TIME_FORMAT)
                log_dict['time_out_display'] = dt_keluar.strftime('%H:%M:%S')
            except:
                log_dict['time_out_display'] = ''
        else:
            log_dict['time_out_display'] = ''
            
        # Hitung Durasi
        durasi_menit = hitung_durasi_menit(log_dict['waktu_masuk'], log_dict['waktu_keluar'])
        if durasi_menit > 0:
            jam = int(durasi_menit // 60)
            menit = int(durasi_menit % 60)
            log_dict['durasi_display'] = f"{jam:02d}:{menit:02d}"
        else:
            log_dict['durasi_display'] = "00:00"
            
        # Hitung Gaji
        gaji = 0
        if durasi_menit > 0:
            gaji = durasi_menit * tarif_per_menit
        
        log_dict['gaji_display'] = format_rupiah(gaji)
            
        processed_logs.append(log_dict)

    return jsonify({'logs': processed_logs})

def proses_item_antrean(key, val, is_silent=False):
    """
    Memproses satu item antrean dengan logika Auto-Silent.
    is_silent: Jika True (dari startup cleanup), paksa silent tanpa kalkulasi umur.
    """
    # 1. Tentukan apakah harus Silent (Tanpa Balasan) berdasarkan Umur Data
    if not is_silent:
        ts_data = val.get('timestamp', 0)
        ts_now = int(datetime.datetime.now().timestamp() * 1000)
        age_seconds = (ts_now - ts_data) / 1000.0
        # Jika data sudah basi > 10 detik, jangan kirim balasan (Silent)
        is_silent = True if age_seconds > 10 else False
    
    prefix_log = "[ONLINE-DELAY]" if is_silent else "[ONLINE]"

    # Fungsi helper internal untuk kirim balasan & hapus queue
    def reply_and_clean(status, msg):
        print(f"  -> {status}: {msg}")
        
        # Hanya tulis respon jika data masih segar
        if not is_silent:
            try:
                db.reference(f'online_responses/{key}').set({
                    'status': status,
                    'msg': msg,
                    'timestamp': {'.sv': 'timestamp'}
                })
            except: pass
        
        try:
            db.reference(f'online_queue/{key}').delete()
        except: pass

    conn = None
    try:
        conn = sqlite3.connect(DB_NAME)
        cur = conn.cursor()
        
        id_kampus = val.get('id_asisten_kampus')
        password_input = val.get('password')
        lab_input = val.get('lokasi_lab')
        kelas_input = val.get('nama_kelas')
        tipe_req = val.get('tipe')
        timestamp_req = val.get('timestamp') 

        print(f"{prefix_log} Request {tipe_req} dari {id_kampus} di {lab_input}...")
        # 1. Validasi Password
        if password_input != ONLINE_TOKEN:
            reply_and_clean('ERROR', 'Password Global Lab Salah')
            return

        # 2. Cek User & Hak Akses
        cur.execute("SELECT fingerprint_id, nama, hak_akses FROM users WHERE id_asisten_kampus=?", (id_kampus,))
        user = cur.fetchone()
        if not user:
            reply_and_clean('ERROR', f'User {id_kampus} tidak ditemukan.')
            return
        
        fp_id, nama, hak_akses_str = user
        
        # 3. Validasi Hak Akses
        list_hak_akses = [h.strip() for h in hak_akses_str.split(',')]
        if lab_input not in list_hak_akses:
            reply_and_clean('ERROR', f'Tidak punya hak akses di {lab_input}.')
            return

        # 4. Validasi Waktu
        waktu_request = datetime.datetime.fromtimestamp(timestamp_req / 1000.0)
        jadwal, wm, ws = get_current_schedule(lab_input, check_time=waktu_request)
        
        if not jadwal or jadwal.get('is_online') != True or jadwal.get('nama_kelas') != kelas_input:
            if tipe_req.upper() == "MASUK":
                reply_and_clean('ERROR', 'Jadwal Kelas/Waktu tidak valid untuk Online.')
                return

        # 5. Eksekusi
        lokasi_final = lab_input 
        kelas_final = kelas_input 
        
        if tipe_req.upper() == "MASUK":
            waktu_masuk_final = max(waktu_request, wm)
            cur.execute("SELECT * FROM active_sessions WHERE fingerprint_id=?", (fp_id,))
            if cur.fetchone():
                reply_and_clean('ERROR', 'Anda sudah login/aktif (Double Login).')
            else:
                cur.execute('''INSERT INTO logs 
                    (fingerprint_id, nama, id_asisten_kampus, waktu_masuk, status, lokasi_lab, kelas, is_synced)
                    VALUES (?, ?, ?, ?, ?, ?, ?, 0)''', 
                    (fp_id, nama, id_kampus, waktu_masuk_final.strftime(DB_TIME_FORMAT), 'MASUK', lokasi_final, kelas_final))
                last_id = cur.lastrowid
                cur.execute('''INSERT INTO active_sessions 
                    (fingerprint_id, nama, id_asisten_kampus, waktu_masuk, jam_selesai_kelas, lokasi_lab, kelas, log_db_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                    (fp_id, nama, id_kampus, waktu_masuk_final.strftime(DB_TIME_FORMAT), ws.strftime(DB_TIME_FORMAT), lokasi_final, kelas_final, last_id))
                reply_and_clean('SUCCESS', f'Halo {nama}, Absen Masuk Berhasil.')

        elif tipe_req.upper() == "KELUAR":
            cur.execute("SELECT log_db_id, jam_selesai_kelas, lokasi_lab, kelas FROM active_sessions WHERE fingerprint_id=?", (fp_id,))
            active = cur.fetchone()
            if active:
                log_db_id, jam_selesai_str, db_lokasi, db_kelas = active
                if db_lokasi != lokasi_final or db_kelas != kelas_final:
                    reply_and_clean('ERROR', f'Data sesi tidak cocok.')
                else:
                    waktu_selesai_db = datetime.datetime.strptime(jam_selesai_str, DB_TIME_FORMAT)
                    waktu_keluar_final = min(waktu_request, waktu_selesai_db)
                    cur.execute("UPDATE logs SET status='Keluar', waktu_keluar=?, is_synced=0 WHERE id=?", 
                               (waktu_keluar_final.strftime(DB_TIME_FORMAT), log_db_id))
                    cur.execute("DELETE FROM active_sessions WHERE fingerprint_id=?", (fp_id,))
                    reply_and_clean('SUCCESS', f'Sampai Jumpa {nama}, Absen Keluar Berhasil.')
            else:
                reply_and_clean('ERROR', 'Tidak ada sesi aktif untuk di-logout.')

        conn.commit()
            
    except Exception as e:
        print(f"[ONLINE PROCESS ERROR] {e}")
    finally:
        if conn:
            conn.close()


def stream_online_queue_listener():
    global IS_MASTER_DATA_READY
    print("[ONLINE-STREAM] Menunggu Master Data siap...")
    while not IS_MASTER_DATA_READY:
        time.sleep(1)
    print("[ONLINE-STREAM] Master Data Siap. Mengecek antrean tertunda...")
    def listener(event):
        if not event.data or event.path == "/" or event.data is None: return
        
        try:
            # Handle Single Item vs Dictionary
            if event.path != "/":
                key = event.path.strip("/")
                val = event.data
                proses_item_antrean(key, val)
            else:
                for key, val in event.data.items():
                    proses_item_antrean(key, val)      
        except Exception as e:
            print(f"[ONLINE LISTENER ERROR] {e}")

    while True:
        print("[ONLINE-STREAM] Memulai/Me-restart Listener Online Queue...")
        
        # 1. Bersihkan respon lama & Proses antrean nyangkut
        try: db.reference('online_responses').delete()
        except: pass
        
        try:
            startup_queue = db.reference('online_queue').get()
            if startup_queue:
                print(f"[ONLINE-STREAM] Memproses {len(startup_queue)} antrean lama secara SILENT...")
                for key, val in startup_queue.items():
                    proses_item_antrean(key, val, is_silent=True)
        except: pass

        # 2. Mulai Listener Baru
        try:
            my_stream = db.reference('online_queue').listen(listener)
            
            # Tahan thread di sini
            while True:
                time.sleep(60)
                
        except Exception as e:
            print(f"[ONLINE-STREAM KONEKSI PUTUS/ERROR] {e}. Me-restart dalam 10 detik...")
            try: my_stream.close()
            except: pass
            time.sleep(10)

def task_queue_watchdog():
    """
    Patroli rutin setiap 30 detik untuk memproses antrean yang 
    mungkin terlewat oleh Listener saat koneksi tidak stabil.
    """
    time.sleep(60)
    print("[ONLINE-WATCHDOG] Patroli Antrean Online Berjalan...")
    while True:
        try:
            queue_data = db.reference('online_queue').get()
            
            if queue_data:
                for key, val in queue_data.items():
                    proses_item_antrean(key, val)
            
        except Exception as e:
            pass
            
        time.sleep(30)

def task_cleanup_firebase_responses():
    """
    Membersihkan data di node '/online_responses' yang sudah basi.
    Cutoff: 15 detik.
    Frekuensi Cek: Setiap 5 menit.
    """
    time.sleep(60)
    print("[CLEANUP-RESPONSES] Thread Pembersihan Respon Firebase Berjalan...")
    while True:
        try:
            # 1. Tentukan batas waktu basi (15 Detik yang lalu)
            cutoff_time = int(datetime.datetime.now().timestamp() * 1000) - 15000
            
            # 2. Query Firebase: Cari data yang 'timestamp' <= cutoff_time
            ref = db.reference('online_responses')
            old_responses = ref.order_by_child('timestamp').end_at(cutoff_time).get()
            
            if old_responses:
                count = 0
                for key, val in old_responses.items():
                    ref.child(key).delete()
                    count += 1
                
                if count > 0:
                    print(f"[CLEANUP-RESPONSES] Berhasil menghapus {count} respon basi (>15 detik).")
                
        except Exception as e:
            print(f"[CLEANUP-RESPONSES ERROR] {e}")
            
        # 3. Tidur selama 5 menit
        time.sleep(300)

def upsert_log_ke_sqlite(cur, fb_key, val):
    """Fungsi bantu untuk Update atau Insert data dari Firebase ke SQLite"""
    
    id_kampus = val.get('id_asisten_kampus', 'Unknown')
    tgl = val.get('tanggal', '').strip()
    t_in = val.get('time_in', '').strip()
    lokasi = val.get('lokasi_lab', '')
    waktu_masuk_str = f"{tgl} {t_in}" if (tgl and t_in) else None

    # --- PERBAIKAN v2: Fallback lookup hanya boleh cocok dengan record lokal yang belum punya firebase_key ---
    # Langkah 1: Cek exact match by Firebase Key
    cur.execute("SELECT id FROM logs WHERE firebase_key = ?", (fb_key,))
    existing = cur.fetchone()
    
    # Langkah 2: Jika tidak ditemukan by key, cek fallback HANYA untuk record lokal (tanpa firebase_key)
    if not existing and waktu_masuk_str:
        cur.execute("""SELECT id FROM logs 
                       WHERE id_asisten_kampus = ? AND waktu_masuk = ? AND lokasi_lab = ? 
                       AND (firebase_key IS NULL OR firebase_key = '')""", 
                    (id_kampus, waktu_masuk_str, lokasi))
        existing = cur.fetchone()
    # ------------------------------------
    
    nama = val.get('nama_asisten', 'Unknown')
    kelas = val.get('kelas', '')
    status = val.get('status', '')
    t_out = val.get('time_out', '').strip()
    waktu_keluar_str = f"{tgl} {t_out}" if (tgl and t_out) else None

    # Validasi format
    try:
        if waktu_masuk_str: datetime.datetime.strptime(waktu_masuk_str, DB_TIME_FORMAT)
        if waktu_keluar_str: datetime.datetime.strptime(waktu_keluar_str, DB_TIME_FORMAT)
    except ValueError:
        print(f"  [SKIP-FORMAT] Data '{fb_key}' dilewati karena format tanggal/jam rusak: tanggal='{tgl}', time_in='{t_in}', time_out='{t_out}'")
        return False # Skip jika format jam rusak
    
    cur.execute("SELECT fingerprint_id FROM users WHERE id_asisten_kampus=?", (id_kampus,))
    user_row = cur.fetchone()
    fp_id = user_row[0] if user_row else 0

    if existing:
        # UPDATE LOGS (Bisa jadi karena Listener mendeteksi key, atau menemukan duplikat waktu)
        log_db_id = existing[0]
        
        # --- PERBAIKAN: Jangan timpa jika data lokal sedang menunggu upload ---
        cur.execute("SELECT is_synced FROM logs WHERE id=?", (log_db_id,))
        sync_row = cur.fetchone()
        if sync_row and sync_row[0] == 0:
            # Data lokal lebih baru (belum ter-upload), SKIP update dari Firebase
            print(f"  [SKIP-UNSYNCED] Data '{fb_key}' ({nama}) dilewati karena data lokal belum ter-upload (is_synced=0).")
            return False
        # -----------------------------------------------------------------
        
        cur.execute('''
            UPDATE logs SET 
            status=?, waktu_masuk=?, waktu_keluar=?, lokasi_lab=?, kelas=?, firebase_key=?, is_synced=1
            WHERE id=?
        ''', (status, waktu_masuk_str, waktu_keluar_str, lokasi, kelas, fb_key, log_db_id))
        
        # UPDATE ACTIVE SESSIONS
        if status.upper() == "MASUK":
            cur.execute('''UPDATE active_sessions SET waktu_masuk=?, lokasi_lab=?, kelas=? WHERE log_db_id=?''', 
                       (waktu_masuk_str, lokasi, kelas, log_db_id))
        elif status.startswith("Keluar"):
            cur.execute("DELETE FROM active_sessions WHERE log_db_id=?", (log_db_id,))
    else:
        # INSERT BARU (Hanya jika benar-benar tidak ada di database sama sekali)
        cur.execute('''
            INSERT INTO logs (firebase_key, fingerprint_id, nama, id_asisten_kampus, waktu_masuk, waktu_keluar, status, lokasi_lab, kelas, is_synced)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
        ''', (fb_key, fp_id, nama, id_kampus, waktu_masuk_str, waktu_keluar_str, status, lokasi, kelas))
        last_log_id = cur.lastrowid
        
        # INSERT KE ACTIVE SESSIONS (Sama seperti sebelumnya)
        if status.upper() == "MASUK" and waktu_masuk_str:
            jam_selesai_str = None
            for jdwl in CACHED_JADWAL:
                lokasi_murni = lokasi.replace(" (ONLINE)", "").strip()
                kelas_murni = kelas.replace(" (ONLINE)", "").strip()
                if jdwl.get('lokasi_lab') == lokasi_murni and jdwl.get('nama_kelas') == kelas_murni:
                    try:
                        masuk_dt = datetime.datetime.strptime(waktu_masuk_str, DB_TIME_FORMAT)
                        js_h, js_m = map(int, jdwl['jam_selesai'].split(':'))
                        selesai_dt = masuk_dt.replace(hour=js_h, minute=js_m, second=0)
                        jam_selesai_str = selesai_dt.strftime(DB_TIME_FORMAT)
                        break
                    except: pass
            
            if not jam_selesai_str:
                masuk_dt = datetime.datetime.strptime(waktu_masuk_str, DB_TIME_FORMAT)
                jam_selesai_str = (masuk_dt + datetime.timedelta(hours=3)).strftime(DB_TIME_FORMAT)
            
            cur.execute("SELECT log_db_id, jam_selesai_kelas, kelas FROM active_sessions WHERE fingerprint_id=?", (fp_id,))
            old_session = cur.fetchone()
            
            if old_session:
                old_log_db_id, old_jam_selesai_str, old_kelas = old_session
                if old_log_db_id != last_log_id:
                    print(f"  ->[DOWN-SYNC] Sesi nyangkut ditemukan untuk {nama}. Auto-Close...")
                    try:
                        old_waktu_keluar_obj = datetime.datetime.strptime(old_jam_selesai_str, DB_TIME_FORMAT)
                        cur.execute('''UPDATE logs SET status='Keluar (Auto-Replaced)', waktu_keluar=?, is_synced=0 WHERE id=?''', 
                                    (old_waktu_keluar_obj.strftime(DB_TIME_FORMAT), old_log_db_id))
                    except: pass
                    cur.execute("DELETE FROM active_sessions WHERE fingerprint_id=?", (fp_id,))
            
            cur.execute('''INSERT INTO active_sessions 
                (fingerprint_id, nama, id_asisten_kampus, waktu_masuk, jam_selesai_kelas, lokasi_lab, kelas, log_db_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)''', (fp_id, nama, id_kampus, waktu_masuk_str, jam_selesai_str, lokasi, kelas, last_log_id))
            print(f"  ->[DOWN-SYNC] Active Session didaftarkan untuk: {nama}")

    return True

def stream_absensi_log_listener():
    """
    Mendengarkan SETIAP perubahan yang terjadi di node 'absensi_log' di Firebase.
    Jika Admin mengubah data (misal: time_in) di Firebase Console, 
    Dashboard Lokal akan langsung terupdate seketika.
    """
    
    def listener(event):
        if event.path == "/" and event.data is None: 
            return
        
        try:
            conn = sqlite3.connect(DB_NAME)
            cur = conn.cursor()

            # KASUS 1: Perubahan pada data spesifik (Admin edit data / Data baru masuk)
            # path akan berisi "/-Oxyz123" atau "/-Oxyz123/time_in"
            if event.path != "/":
                # Ambil Firebase Key-nya (Bagian pertama dari path)
                path_parts = event.path.strip("/").split("/")
                fb_key = path_parts[0]

                if event.data is None:
                    # --- FITUR BARU: SINKRONISASI PENGHAPUSAN ---
                    # Jika data di Firebase dihapus, hapus juga dari SQLite
                    cur.execute("DELETE FROM logs WHERE firebase_key=?", (fb_key,))
                    print(f"[LOG-STREAM] Data dihapus Admin di Firebase. Menghapus log lokal: {fb_key}")
                else:
                    # Data diupdate / Data baru
                    log_data = db.reference(f'absensi_log/{fb_key}').get()
                    if log_data:
                        upsert_log_ke_sqlite(cur, fb_key, log_data)
                        print(f"[LOG-STREAM] Data sinkron: {log_data.get('nama_asisten')} | Status: {log_data.get('status')}")

            # KASUS 2: Initial Load (Saat Server Baru Nyala)
            # Firebase akan mengirim seluruh isi database. Kita ambil 50 terbaru saja agar PC tidak lag.
            else:
                if event.data:
                    print("[LOG-STREAM] Memuat data awal dari Firebase...")
                    # Kita yang melakukan filter 50 terakhir secara manual di Python
                    sorted_logs = sorted(event.data.items(), key=lambda item: item[1].get('timestamp_ms', 0), reverse=True)[:50]
                    #Initial Load Untuk Mengambil Semua Data Lama absensi_log di Firebase
                    # def sort_by_event_time(item):
                    #     val = item[1]
                    #     tgl = val.get('tanggal', '').strip()
                    #     t_in = val.get('time_in', '').strip()
                    #     t_out = val.get('time_out', '').strip()
                        
                    #     if tgl and t_in:
                    #         return f"{tgl} {t_in}"
                    #     elif tgl and t_out:
                    #         return f"{tgl} {t_out}"
                    #     else:
                    #         # Jika data korup, taruh di urutan paling bawah
                    #         return "9999-12-31 23:59:59"
                    # sorted_logs = sorted(event.data.items(), key=sort_by_event_time, reverse=False)
                    count = 0
                    skipped = 0
                    for fb_key, log_data in sorted_logs:
                        if upsert_log_ke_sqlite(cur, fb_key, log_data):
                            count += 1
                        else:
                            skipped += 1
                    print(f"[LOG-STREAM] Selesai memuat {count} data. ({skipped} dilewati dari total {count + skipped})")

            conn.commit()
            conn.close()

        except Exception as e:
            print(f"[LOG-STREAM ERROR] {e}")

    # --- LOGIKA SELF-HEALING (INFINITE LOOP) ---
    while True:
        print("[LOG-STREAM] Memulai/Me-restart Listener Log Absensi...")
        try:
            # Mulai mendengarkan
            my_stream = db.reference('absensi_log').listen(listener)
            
            # Loop penahan agar thread tidak selesai.
            # Jika my_stream putus/crash dari sisi server, ini akan memicu Exception
            while True:
                time.sleep(60) 
                
        except Exception as e:
            print(f"[LOG-STREAM KONEKSI PUTUS/ERROR] {e}. Me-restart Listener dalam 10 detik...")
            # Tutup stream lama jika masih ada sisa-sisa
            try: my_stream.close() 
            except: pass
            
            # Tunggu sebentar sebelum mencoba menyambung lagi
            time.sleep(10)

def task_absensi_log_watchdog():
    """
    Patroli rutin setiap 10 menit (600 detik) untuk memastikan 
    tidak ada data edit dari Admin di Firebase yang terlewat oleh Listener.
    """
    time.sleep(60)
    print("[LOG-WATCHDOG] Patroli Downstream Sync Berjalan...")
    while True:
        try:
            # Ambil 50 data terakhir dari Firebase (Sama seperti listener saat startup)
            fb_data = db.reference('absensi_log').order_by_key().limit_to_last(50).get()
            
            if fb_data:
                conn = sqlite3.connect(DB_NAME)
                try:
                    cur = conn.cursor()
                    for fb_key, val in fb_data.items():
                        # Panggil fungsi helper yang sudah kita buat sebelumnya
                        # Fungsi ini aman dipanggil berulang kali (Idempotent)
                        upsert_log_ke_sqlite(cur, fb_key, val)
                        
                    conn.commit()
                finally:
                    conn.close()
                    
        except Exception as e:
            # Error wajar jika internet sedang mati saat patroli
            pass
            
        # Tidur selama 10 menit (600 detik)
        # Tidak perlu sering-sering karena ini hanya backup untuk Listener
        time.sleep(600)

def safe_delete_firebase_node(path):
    """
    Menghapus node Firebase secara background. 
    Dilengkapi dengan fitur Retry agar tidak gagal jika internet sedang mati.
    """
    while True:
        try:
            db.reference(path).delete()
            print(f"[CLEANUP] Berhasil menghapus node lama di Firebase: {path}")
            break
        except Exception:
            time.sleep(10) # Jika offline, tunggu 10 detik lalu coba lagi

if __name__ == '__main__':
    init_db()
    load_jadwal_lokal()

    # Threads Utama
    threading.Thread(target=task_smart_auto_tap_out, daemon=True).start()
    threading.Thread(target=task_upload_logs, daemon=True).start()
    threading.Thread(target=task_upload_master_data, daemon=True).start()

    # Threads Listener (Real-Time)
    threading.Thread(target=stream_master_data_listener, daemon=True).start()
    threading.Thread(target=stream_device_control_listener, daemon=True).start()
    threading.Thread(target=stream_online_queue_listener, daemon=True).start()
    threading.Thread(target=stream_absensi_log_listener, daemon=True).start()

    # Threads Watchdog (Patroli Backup)
    threading.Thread(target=task_queue_watchdog, daemon=True).start()
    threading.Thread(target=task_absensi_log_watchdog, daemon=True).start()
    threading.Thread(target=task_master_data_watchdog, daemon=True).start()
    threading.Thread(target=task_cleanup_firebase_responses, daemon=True).start()
    
    print("SERVER BERJALAN di Port 5000...")
    app.run(host='0.0.0.0', port=5000, debug=False)