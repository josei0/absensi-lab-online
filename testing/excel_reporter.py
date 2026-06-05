"""
SmartLab IoT Testing Framework — Excel Report Engine
=====================================================
Menggunakan openpyxl untuk menghasilkan file .xlsx dengan styling profesional:
- Header merah maroon (#CC0000) dengan font putih bold
- Auto-column-width
- Conditional formatting (PASS=hijau, WARNING=kuning, FAIL=merah)
- Summary rows otomatis
- Metadata sheet (tanggal, tester, environment)
"""

import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from testing.config import (
    EXCEL_HEADER_COLOR, EXCEL_HEADER_FONT_COLOR,
    EXCEL_PASS_COLOR, EXCEL_WARNING_COLOR, EXCEL_FAIL_COLOR,
    RESULTS_DIR
)
from testing.test_helpers import ensure_results_dir


# ================================================================
# STYLE DEFINITIONS
# ================================================================

# Header style
HEADER_FILL = PatternFill(start_color=EXCEL_HEADER_COLOR, end_color=EXCEL_HEADER_COLOR, fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=EXCEL_HEADER_FONT_COLOR)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Data cell style
DATA_FONT = Font(name="Calibri", size=10)
DATA_ALIGNMENT = Alignment(horizontal="center", vertical="center")
DATA_ALIGNMENT_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Conditional fills
PASS_FILL = PatternFill(start_color=EXCEL_PASS_COLOR, end_color=EXCEL_PASS_COLOR, fill_type="solid")
WARNING_FILL = PatternFill(start_color=EXCEL_WARNING_COLOR, end_color=EXCEL_WARNING_COLOR, fill_type="solid")
FAIL_FILL = PatternFill(start_color=EXCEL_FAIL_COLOR, end_color=EXCEL_FAIL_COLOR, fill_type="solid")

# Border
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Summary row style
SUMMARY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
SUMMARY_FONT = Font(name="Calibri", size=10, bold=True)


# ================================================================
# EXCEL REPORTER CLASS
# ================================================================

class ExcelReporter:
    """
    Engine untuk membuat file Excel report terformat cantik.
    
    Usage:
        reporter = ExcelReporter("01_Fingerprint_Accuracy")
        reporter.add_sheet("Akurasi Fingerprint", headers, data_rows)
        reporter.add_sheet("Detail Percobaan", headers2, data_rows2)
        filepath = reporter.save()
    """

    def __init__(self, filename_prefix, metadata=None):
        """
        Args:
            filename_prefix: Nama file tanpa tanggal dan ekstensi.
                             Contoh: "01_Fingerprint_Accuracy"
            metadata: dict opsional berisi info pengujian
                      {"tester": "...", "environment": "...", "notes": "..."}
        """
        self.filename_prefix = filename_prefix
        self.metadata = metadata or {}
        self.wb = Workbook()
        # Hapus sheet default
        self.wb.remove(self.wb.active)
        self._sheet_count = 0

    def add_sheet(self, sheet_name, headers, data_rows, status_col_name=None,
                  left_align_cols=None, summary_row=None):
        """
        Tambahkan sheet baru ke workbook.
        
        Args:
            sheet_name: Nama sheet (max 31 karakter untuk Excel compatibility)
            headers: list[str] — nama kolom header
            data_rows: list[list] — baris data (setiap baris = list values)
            status_col_name: str — nama kolom yang berisi PASS/FAIL/WARNING
                             untuk conditional formatting. None jika tidak ada.
            left_align_cols: list[str] — nama kolom yang di-align kiri (untuk teks panjang)
            summary_row: list — baris summary opsional (total/rata-rata)
        """
        # Truncate sheet name jika >31 karakter
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]

        ws = self.wb.create_sheet(title=sheet_name)
        self._sheet_count += 1

        if left_align_cols is None:
            left_align_cols = []

        # --- WRITE HEADERS ---
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        # --- WRITE DATA ROWS ---
        status_col_idx = None
        if status_col_name and status_col_name in headers:
            status_col_idx = headers.index(status_col_name) + 1

        for row_idx, row_data in enumerate(data_rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER

                # Alignment
                header_name = headers[col_idx - 1] if col_idx <= len(headers) else ""
                if header_name in left_align_cols:
                    cell.alignment = DATA_ALIGNMENT_LEFT
                else:
                    cell.alignment = DATA_ALIGNMENT

            # Conditional formatting pada baris berdasarkan kolom status
            if status_col_idx and status_col_idx <= len(row_data):
                status_val = str(row_data[status_col_idx - 1]).upper().strip()
                fill = None
                if status_val in ("PASS", "SUCCESS", "ONLINE", "YES", "EXCELLENT"):
                    fill = PASS_FILL
                elif status_val in ("WARNING", "PARTIAL"):
                    fill = WARNING_FILL
                elif status_val in ("FAIL", "FAILED", "OFFLINE", "NO", "ERROR"):
                    fill = FAIL_FILL

                if fill:
                    for col_idx in range(1, len(row_data) + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill

        # --- WRITE SUMMARY ROW ---
        if summary_row:
            summary_row_idx = len(data_rows) + 2
            for col_idx, value in enumerate(summary_row, start=1):
                cell = ws.cell(row=summary_row_idx, column=col_idx, value=value)
                cell.font = SUMMARY_FONT
                cell.fill = SUMMARY_FILL
                cell.alignment = DATA_ALIGNMENT
                cell.border = THIN_BORDER

        # --- AUTO COLUMN WIDTH ---
        self._auto_column_width(ws, headers, data_rows)

        # Freeze header row
        ws.freeze_panes = "A2"

        return ws

    def add_metadata_sheet(self):
        """Tambahkan sheet 'Info Pengujian' di posisi pertama."""
        ws = self.wb.create_sheet(title="Info Pengujian", index=0)

        meta_rows = [
            ["Informasi Pengujian", ""],
            ["", ""],
            ["Tanggal Pengujian", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Tester", self.metadata.get("tester", "SmartLab Testing Framework")],
            ["Environment", self.metadata.get("environment", "Test Server (Port 5001)")],
            ["Database", self.metadata.get("database", "test_database_lab.sqlite (Isolated)")],
            ["Firebase Node", self.metadata.get("firebase_node", "/test_sandbox/")],
        ]

        # Notes
        if self.metadata.get("notes"):
            meta_rows.append(["", ""])
            meta_rows.append(["Catatan", self.metadata["notes"]])

        for row_idx, (label, value) in enumerate(meta_rows, start=1):
            cell_label = ws.cell(row=row_idx, column=1, value=label)
            cell_value = ws.cell(row=row_idx, column=2, value=value)

            if row_idx == 1:
                cell_label.font = Font(name="Calibri", size=14, bold=True, color=EXCEL_HEADER_COLOR)
            elif label:
                cell_label.font = Font(name="Calibri", size=10, bold=True)
                cell_value.font = DATA_FONT

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 50

    def save(self):
        """
        Simpan workbook ke folder results/.
        Returns: path lengkap file yang disimpan.
        """
        ensure_results_dir()

        date_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{self.filename_prefix}_{date_str}.xlsx"
        filepath = os.path.join(RESULTS_DIR, filename)

        # Tambahkan metadata sheet jika belum ada
        if "Info Pengujian" not in self.wb.sheetnames:
            self.add_metadata_sheet()

        self.wb.save(filepath)
        print(f"[EXCEL] Laporan disimpan: {filepath}")
        return filepath

    # --- INTERNAL HELPERS ---

    @staticmethod
    def _auto_column_width(ws, headers, data_rows):
        """Otomatis menyesuaikan lebar kolom berdasarkan konten."""
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_length = len(str(header)) + 2  # header length + padding

            for row_data in data_rows:
                if col_idx <= len(row_data):
                    cell_value = str(row_data[col_idx - 1]) if row_data[col_idx - 1] is not None else ""
                    max_length = max(max_length, len(cell_value) + 2)

            # Cap at 50 characters width
            ws.column_dimensions[col_letter].width = min(max_length, 50)
